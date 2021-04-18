# -*- coding: utf-8 -*-
"""
Created on Wed May 13 13:59:31 2020

@author: Bernhard Foellmer
"""

import sys, os
sys.path.append('H:/cloud/cloud_data/Projects/DL/Code/src')
sys.path.append('H:/cloud/cloud_data/Projects/DL/Code/src/ct')
import pandas as pd
import ntpath
import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import Font, Color, Border, Side
from openpyxl.styles import Protection
from openpyxl.styles import PatternFill
from glob import glob
from shutil import copyfile
import numpy as np
from collections import defaultdict
from openpyxl.utils import get_column_letter
from CTDataStruct import CTPatient
import keyboard
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from settings import initSettings, saveSettings, loadSettings, fillSettingsTags
from classification import createRFClassification, initRFClassification, classifieRFClassification
from filterTenStepsGuide import filter_CACS_10StepsGuide, filter_CACS, filter_NCS, filterReconstruction, filter_CTA, filer10StepsGuide, filterReconstructionRF
from discharge_extract import extractDICOMTags
from tqdm import tqdm
#from reco.reco_filter import RecoFilter

patient_status = ['OK', 'EXCLUDED', 'MISSING_CACS', 'MISSING_CTA', 'MISSING_NC_CACS', 'MISSING_NC_CTA']
patient_status_manual = ['OK', 'EXCLUDED', 'UNDEFINED', 'INPROGRESS']
patient_status_manualStr = '"' + 'OK,' + 'EXCLUDED,' + 'UNDEFINED,' + 'INPROGRESS,' + '"'

scanClasses = defaultdict(lambda:None,{'UNDEFINED': 0, 'CACS': 1, 'CTA': 2, 'NCS_CACS': 3, 'NCS_CTA': 4, 'ICA': 5, 'OTHER': 6})
scanClassesInv = defaultdict(lambda:None,{0: 'UNDEFINED', 1: 'CACS', 2: 'CTA', 3: 'NCS_CACS', 4: 'NCS_CTA', 5: 'ICA', 6: 'OTHER'})
scanClassesStr = '"' + 'UNDEFINED,' + 'CACS,' + 'CTA,' + 'NCS_CACS,' + 'NCS_CTA,' + 'ICA,' + 'OTHER' +'"'
scanClassesManualStr = '"' + 'UNDEFINED,' + 'CACS,' + 'CTA,' + 'NCS_CACS,' + 'NCS_CTA,' + 'ICA,' + 'OTHER,' + 'PROBLEM,' + 'QUESTION,' +'"'
imageQualityStr = '"' + 'UNDEFINED,' + 'GOOD,' + 'BAD' +'"'
recoClasses = ['FBP', 'IR', 'UNDEFINED']
changeClasses = ['NO_CHANGE', 'SOURCE_CHANGE', 'MASTER_CHANGE', 'MASTER_SOURCE_CHANGE']

def setColor(workbook, sheet, rows, NumColumns, color):
    for r in rows:
        if r % 100 == 0:
            print('index:', r, '/', max(rows))
        for c in range(1,NumColumns):
            cell = sheet.cell(r, c)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type = 'solid')

def setColorFormula(sheet, formula, color, NumRows, NumColumns):
    column_letter = get_column_letter(NumColumns+1)
    colorrange="B2:" + str(column_letter) + str(NumRows)
    dxf = DifferentialStyle(font=Font(color=color))
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = [formula]
    sheet.conditional_formatting.add(colorrange, r)

def setBorderFormula(sheet, formula, NumRows, NumColumns):
    column_letter = get_column_letter(NumColumns+1)
    colorrange="B1:" + str(column_letter) + str(NumRows)
    thin = Side(border_style="thin", color="000000")
    border = Border(bottom=thin)
    dxf = DifferentialStyle(border=border)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = [formula]
    sheet.conditional_formatting.add(colorrange, r)
        
    # Set border for index
    for i in range(1, NumRows + 1):
        cell = sheet.cell(i, 1)
        cell.border = Border()
    return sheet
        
def sortFilepath(filepathList):
    filenameList=[]
    folderpathList=[]
    for filepath in filepathList:
        folderpath, filename, _ = splitFilePath(filepath)
        filenameList.append(filename)
        folderpathList.append(folderpath)
        
    dates_str = [x.split('_')[-1] for x in filenameList]
    dates = [datetime.datetime(int(x[4:8]), int(x[2:4]), int(x[0:2])) for x in dates_str]
    idx = list(np.argsort(dates))
    filepathlistsort=[]
    for i in idx:
        filepathlistsort.append(folderpathList[i] + '/' + '_'.join(filenameList[i].split('_')[0:-1]) + '_' + dates[i].strftime("%d%m%Y") + '.xlsx')
    
    return filepathlistsort

def sortFolderpath(folderpath, folderpathList):
    dates_str = [x.split('_')[-1] for x in folderpathList]
    dates = [datetime(int(x[4:8]), int(x[2:4]), int(x[0:2])) for x in dates_str]
    date_str = folderpath.split('_')[-1]
    date = datetime(int(date_str[4:8]), int(date_str[2:4]), int(date_str[0:2]))
    idx = list(np.argsort(dates))
    folderpathSort=[]
    for i in idx:
        folderpathSort.append(folderpathList[i])
        if dates[i] == date:
            break
    return folderpathSort

def isNaN(num):
    return num != num

def splitFilePath(filepath):
    """ Split filepath into folderpath, filename and file extension
    
    :param filepath: Filepath
    :type filepath: str
    """
    folderpath, _ = ntpath.split(filepath)
    head, file_extension = os.path.splitext(filepath)
    folderpath, filename = ntpath.split(head)
    return folderpath, filename, file_extension

 

def update_CACS_10StepsGuide(df_CACS, sheet):
    for index, row in df_CACS.iterrows():
        cell_str = 'AB' + str(index+2)
        cell = sheet[cell_str]
        cell.value = row['CACS10StepsGuide']
        #cell.protection = Protection(locked=False)
    return sheet    

def mergeITT(df_ITT, df_data):
    # Merge ITT table   
    print('Merge ITT table')
    for i in range(len(df_data)):
        patient = df_ITT[df_ITT['ID']==df_data.loc[i, 'PatientID']]
        if len(patient)==1:
            df_data.loc[i, 'ITT'] = patient.iloc[0]['ITT']
            df_data.loc[i, 'Date CT'] = patient.iloc[0]['Date CT']
            df_data.loc[i, 'Date ICA'] = patient.iloc[0]['Date ICA']
    return df_data

def mergeDicom(df_dicom, df_data_old=None):
    print('Merge dicom table')
    if df_data_old is None:
        df_data = df_dicom.copy()
    else:
        idx = df_dicom['SeriesInstanceUID'].isin(df_data_old['SeriesInstanceUID'])
        df_data = pd.concat([df_data_old, df_dicom[idx==False]], axis=0)     
    return df_data

def mergeTracking(df_tracking, df_data, df_data_old=None):
    
    if df_data_old is None:
        
        df_data = df_data.copy()
        df_tracking = df_tracking.copy()
        df_data.replace(to_replace=[np.nan], value='', inplace=True)
        df_tracking.replace(to_replace=[np.nan], value='', inplace=True)
        
        # Merge tracking table
        print('Merge tracking table')
        df_data['Responsible Person Problem'] = ''
        df_data['Date Query'] = ''
        df_data['Date Answer'] = ''
        df_data['Problem Summary'] = ''
        df_data['Results'] = ''
        for index, row in df_tracking.iterrows():
            patient = row['PatientID']
            df_patient = df_data[df_data['PatientID']==patient]
            for indexP, rowP in df_patient.iterrows(): 
                # Update 'Problem Summary'
                if df_data.loc[indexP, 'Problem Summary']=='':
                    df_data.loc[indexP, 'Problem Summary'] = row['Problem Summary']
                else:
                    df_data.loc[indexP, 'Problem Summary'] = df_data.loc[indexP, 'Problem Summary'] + ' | ' + row['Problem Summary']
                # Update 'results'
                if df_data.loc[indexP, 'Results']=='':
                    df_data.loc[indexP, 'Results'] = row['results']
                else:
                    df_data.loc[indexP, 'Results'] = df_data.loc[indexP, 'Results'] + ' | ' + row['results']
    else:
        
        df_data = df_data.copy()
        df_data_old = df_data_old.copy()
        df_tracking = df_tracking.copy()
        df_data.replace(to_replace=[np.nan], value='', inplace=True)
        df_data_old.replace(to_replace=[np.nan], value='', inplace=True)
        df_tracking.replace(to_replace=[np.nan], value='', inplace=True)
    
        l = len(df_data_old)
        df_data['Responsible Person Problem'] = ''
        df_data['Date Query'] = ''
        df_data['Date Answer'] = ''
        df_data['Problem Summary'] = ''
        df_data['Results'] = ''
        df_data['Responsible Person Problem'][0:l] = df_data_old['Responsible Person Problem']
        df_data['Date Query'][0:l] = df_data_old['Date Query']
        df_data['Date Answer'][0:l] = df_data_old['Date Answer']
        df_data['Problem Summary'][0:l] = df_data_old['Problem Summary']
        df_data['Results'][0:l] = df_data_old['Results']
        
        for index, row in df_tracking.iterrows():
            patient = row['PatientID']
            df_patient = df_data[df_data['PatientID']==patient]
            for indexP, rowP in df_patient.iterrows():
                # Update 'Problem Summary'
                if df_data.loc[indexP, 'Problem Summary']=='':
                    df_data.loc[indexP, 'Problem Summary'] = row['Problem Summary']
                else:
                    if not row['Problem Summary'] in df_data.loc[indexP, 'Problem Summary']:
                        df_data.loc[indexP, 'Problem Summary'] = df_data.loc[indexP, 'Problem Summary'] + ' | ' + row['Problem Summary']
                # Update 'results'
                if df_data.loc[indexP, 'Results']=='':
                    df_data.loc[indexP, 'Results'] = row['results']
                else:
                    if not row['results'] in df_data.loc[indexP, 'Results']:
                        df_data.loc[indexP, 'Results'] = df_data.loc[indexP, 'Results'] + ' | ' + row['results']
    return df_data

def mergeEcrf(df_ecrf, df_data):
    
    # Merge ecrf table
    print('Merge ecrf table')
    df_data['1. Date of CT scan'] = ''
    for index, row in df_ecrf.iterrows():
        patient = row['Patient identifier']
        df_patient = df_data[df_data['PatientID']==patient]
        for indexP, rowP in df_patient.iterrows(): 
            # Update '1. Date of CT scan'
            df_data.loc[indexP, '1. Date of CT scan'] = row['1. Date of CT scan']

    return df_data

def mergePhase_exclude_stenosis(df_phase_exclude_stenosis, df_data):
    # Merge phase_exclude_stenosis
    print('Merge phase_exclude_stenosis table')
    df_data['phase_i0011'] = ''
    df_data['phase_i0012'] = ''
    for index, row in df_phase_exclude_stenosis.iterrows():
        patient = row['mnpaid']
        df_patient = df_data[df_data['PatientID']==patient]
        for indexP, rowP in df_patient.iterrows(): 
            # Update tags
            if df_data.loc[indexP, 'phase_i0011']=='':
                df_data.loc[indexP, 'phase_i0011'] = str(row['phase_i0011'])
            else:
                df_data.loc[indexP, 'phase_i0011'] = str(df_data.loc[indexP, 'phase_i0011']) + ', ' + str(row['phase_i0011'])
                
            if df_data.loc[indexP, 'phase_i0012']=='':
                df_data.loc[indexP, 'phase_i0012'] = str(row['phase_i0012'])
            else:
                df_data.loc[indexP, 'phase_i0012'] = str(df_data.loc[indexP, 'phase_i0011']) + ', ' + str(row['phase_i0011'])

    return df_data

def mergePrct(df_prct, df_data):
    # Merge phase_exclude_stenosis
    print('Merge prct table')
    df_data['other_best_phase'] = ''
    df_data['rca_best_phase'] = ''
    df_data['lad_best_phase'] = ''
    df_data['lcx_best_phase'] = ''
    for index, row in df_prct.iterrows():
        patient = row['PatientId']
        df_patient = df_data[df_data['PatientID']==patient]
        for indexP, rowP in df_patient.iterrows(): 
            # Update tags
            df_data.loc[indexP, 'other_best_phase'] = row['other_best_phase']
            df_data.loc[indexP, 'rca_best_phase'] = row['rca_best_phase']
            df_data.loc[indexP, 'lad_best_phase'] = row['lad_best_phase']
            df_data.loc[indexP, 'lcx_best_phase'] = row['lcx_best_phase']
    return df_data


def mergeStenosis_bigger_20_phase(df_stenosis_bigger_20_phases, df_data):
    # Merge phase_exclude_stenosis
    print('Merge Stenosis_bigger_20_phase table')
    df_data['STENOSIS'] = ''
    
    patientnames = df_stenosis_bigger_20_phases['mnpaid'].unique()
    df_stenosis_bigger_20_phases.replace(to_replace=[np.nan], value='', inplace=True)
    for patient in patientnames:
        patientStenose = df_stenosis_bigger_20_phases[df_stenosis_bigger_20_phases['mnpaid']==patient]
        sten = ''
        for index, row in patientStenose.iterrows(): 
            art=''
            if row['LAD']==1:
                art = 'LAD'
            if row['RCA']==1:
                art = 'RCA'
            if row['LMA']==1:
                art = 'LMA'
            if row['LCX']==1:
                art = 'LCX'
            if sten =='':
                if not art=='':
                    sten = art + ':' + str(row['sten_i0231 (Phase #1)']) + ':' + str(row['sten_i0241']) + ':' + str(row['sten_i0251'])
            else:
                if not art=='':
                    sten = sten + ', ' + art + ':' + str(row['sten_i0231 (Phase #1)']) + ':' + str(row['sten_i0241']) + ':' + str(row['sten_i0251'])
        
        df_patient = df_data[df_data['PatientID']==patient]
        for indexP, rowP in df_patient.iterrows(): 
            df_data.loc[indexP, 'STENOSIS'] = sten

    return df_data


def freeze(writer, sheetname, df):
    NumRows=1
    NumCols=1
    df.to_excel(writer, sheet_name = sheetname, freeze_panes = (NumCols, NumRows))
 

def highlight_columns(sheet, columns=[], color='A5A5A5', offset=2):
    for col in columns:
        cell = sheet.cell(1, col+offset)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type = 'solid')
    return sheet  

def setAccessRights(sheet, columns=[], promt='', promptTitle='', formula1='"Dog,Cat,Bat"'):
    for column in columns:
        column_letter = get_column_letter(column+2)
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        dv.prompt = promt
        dv.promptTitle = promptTitle
        column_str = column_letter + str(1) + ':' + column_letter + str(1048576)
        dv.add(column_str)
        sheet.add_data_validation(dv)
    return sheet 

def setComment(sheet, columns=[], comment=''):
    for column in columns:
        column_letter = get_column_letter(column+2)
        dv = DataValidation()
        dv.prompt = comment
        column_str = column_letter + str(1) + ':' + column_letter + str(1048576)
        dv.add(column_str)
        sheet.add_data_validation(dv)
    return sheet  

def checkTables(settings):
    print('Checking existance of required tables.')
    # Check if requird tables exist
    tables=['filepath_dicom', 'filepath_ITT', 'filepath_ecrf', 'filepath_prct',
            'filepath_phase_exclude_stenosis', 'filepath_stenosis_bigger_20_phases', 'filepath_tracking']
    for table in tables:
        if not os.path.isfile(settings[table]):
            raise ValueError("Source file " + settings[table] + ' does not exist. Please copy file in the correct directory!')
    return True

def createData(settings, NumSamples=None):
    """ Create data columns from dicom metadata
        
    :param settings: Dictionary of settings
    :type settings: dict
    """   
    
    XA=False
    # Extract dicom data
    df_dicom = pd.read_excel(settings['filepath_dicom'], index_col=0)
    
    # Reorder datafame
    df_dicom = df_dicom[settings['dicom_tags_order']]
    if XA:
        df_dicom = df_dicom[(df_dicom['Modality']=='CT') | (df_dicom['Modality']=='OT') | (df_dicom['Modality']=='XA')]
    else:
        df_dicom = df_dicom[(df_dicom['Modality']=='CT') | (df_dicom['Modality']=='OT')]
    df_dicom = df_dicom.reset_index(drop=True)
    cols = df_dicom.columns.tolist()
    cols_new = settings['dicom_tags_first'] + [x for x in cols if x not in settings['dicom_tags_first']]
    df_dicom = df_dicom[cols_new]
    df_data = df_dicom.copy()
    df_data = df_data.reset_index(drop=True)
    
    if NumSamples is not None:
        df_data = df_data[0:NumSamples]

    # Extract ecrf data
    df_ecrf = pd.read_excel(settings['filepath_ecrf'])
    df_data = mergeEcrf(df_ecrf, df_data)
    
    # Extract ITT 
    df_ITT = pd.read_excel(settings['filepath_ITT'], 'Tabelle1')
    df_data = mergeITT(df_ITT, df_data)
    
    # Extract phase_exclude_stenosis 
    df_phase_exclude_stenosis = pd.read_excel(settings['filepath_phase_exclude_stenosis'])
    df_data = mergePhase_exclude_stenosis(df_phase_exclude_stenosis, df_data)
    
    # Extract prct
    df_prct = pd.read_excel(settings['filepath_prct'])
    df_data = mergePrct(df_prct, df_data)
    
    # Extract stenosis_bigger_20_phases
    df_stenosis_bigger_20_phases = pd.read_excel(settings['filepath_stenosis_bigger_20_phases'])
    df_data = mergeStenosis_bigger_20_phase(df_stenosis_bigger_20_phases, df_data)  
    
    # Reoder columns
    cols = df_data.columns.tolist()
    cols_new = settings['dicom_tags_first'] + [x for x in cols if x not in settings['dicom_tags_first']]
    
    #filepath_master_data = os.path.join(settings['folderpath_components'], 'discharge_master_data_' + settings['date'] + '.xlsx')
    #df_data.to_excel(settings['filepath_data'])
    df_data.to_pickle(settings['filepath_data'])
    


def createPredictions(settings):
    """ Create prediction columns
        
    :param settings: Dictionary of settings
    :type settings: dict
    """   
    
    df_data = pd.read_pickle(settings['filepath_data'])
    df_pred = pd.DataFrame()
    
    # Filter by CACS based on 10-Steps-Guide
    df = filter_CACS_10StepsGuide(df_data)
    df_pred['CACS10StepsGuide'] = df['CACS10StepsGuide']
    
    # Filter by CACS based  selection
    df = filter_CACS(df_data)
    df_pred['CACS'] = df['CACS']
    
    # Filter by NCS_CACS and NCS_CTA based on  criteria
    df = filter_NCS(df_data)
    df_pred['NCS_CTA'] = df['NCS_CTA']
    df_pred['NCS_CACS'] = df['NCS_CACS']
    
    # Filter by CTA
    df = filter_CTA(settings)
    df_pred['CTA'] = df['CTA'].astype('bool')
    df_pred['CTA_phase'] = df['phase']
    df_pred['CTA_arteries'] = df['arteries']
    df_pred['CTA_source'] = df['source']
   
    # Filter by ICA
    df = pd.DataFrame('', index=np.arange(len(df_pred)), columns=['ICA'])
    df_pred['ICA'] = df['ICA']
 
    # Filter by reconstruction
    df = filterReconstruction(df_data, settings)
    df_pred['RECO'] = df['RECO']

    # Predict CLASS
    classes = ['CACS', 'CTA', 'NCS_CTA', 'NCS_CACS']
    for i in range(len(df_pred)):
        if i % 1000 == 0:
            print('index:', i, '/', len(df_pred))
        value=''
        for c in classes:
            if df_pred.loc[i, c]:
                if value=='':
                    value = value + c
                else:
                    value = value + '+' + c
        if value == '':
            value = 'UNDEFINED'
        df_pred.loc[i, 'CLASS'] = value
        
    # Save predictions    
    df_pred.to_pickle(settings['filepath_prediction'])

def updateRFClassification(folderpath_master, folderpath_master_before):
    """ Update random forest classification
        
    :param settings: Dictionary of settings
    :type settings: dict
    """ 
    date = folderpath_master.split('_')[-1]
    folderpath_components = os.path.join(folderpath_master, 'discharge_components_' + date)
    filepath_rfc = os.path.join(folderpath_components, 'discharge_rfc_' + date + '.xlsx')
    
    folderpath_master_before_list = glob(folderpath_master_before + '/*master*')
    folderpath_master_before_list = sortFolderpath(folderpath_master, folderpath_master_before_list)
    filepathMasters = glob(folderpath_master_before_list[-2] + '/*process*.xlsx')
    
    
    date_before = folderpath_master_before_list[-2].split('_')[-1]
    df_master = pd.read_excel(filepathMasters[0], sheet_name='MASTER_' + date_before)
    columns = ['RFCLabel', 'RFCClass', 'RFCConfidence']
    df_rfc = pd.DataFrame('UNDEFINED', index=np.arange(len(df_master)), columns=columns)
    df_rfc[columns] = df_master[columns]

    df_rfc.to_excel(filepath_rfc)


def createManualSelection(settings):
    """ Create manual selection columns
        
    :param settings: Dictionary of settings
    :type settings: dict
    """ 
    print('Create manual selection')
    #df_data = pd.read_excel(settings['filepath_data'], index_col=0)
    df_data = pd.read_pickle(settings['filepath_data'])
    df_manual0 = pd.DataFrame('UNDEFINED', index=np.arange(len(df_data)), columns=['ClassManualCorrection'])
    df_manual1 = pd.DataFrame('', index=np.arange(len(df_data)), columns=['Comment'])
    df_manual2 = pd.DataFrame('', index=np.arange(len(df_data)), columns=['Responsible Person'])
    df_manual3 = pd.DataFrame('UNDEFINED', index=np.arange(len(df_data)), columns=['Image Quality'])
    df_manual = pd.concat([df_manual0, df_manual1, df_manual2, df_manual3], axis=1)
    #df_manual.to_excel(settings['filepath_manual'])
    df_manual.to_pickle(settings['filepath_manual'])
    
def createTrackingTable(settings):
    """ Create tracking table
        
    :param settings: Dictionary of settings
    :type settings: dict
    """   
    
    print('Create tracking table')
    df_track = pd.DataFrame(columns=settings['columns_tracking'])
    df_track.to_pickle(settings['filepath_master_track'])
    # Update master
    writer = pd.ExcelWriter(settings['filepath_master'], engine="openpyxl", mode="a")
    # Remove sheet if already exist
    sheet_name = 'TRACKING' + '_' + settings['date']
    workbook  = writer.book
    sheetnames = workbook.sheetnames
    if sheet_name in sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
        
    # Add patient ro master
    df_track.to_excel(writer, sheet_name=sheet_name)
    writer.save()

    print('Update tracking table')
    # Read tracking table
    df_tracking = pd.read_excel(settings['filepath_tracking'], 'tracking table')
    df_tracking.replace(to_replace=[np.nan], value='', inplace=True)
    df_track = pd.read_excel(settings['filepath_master'], 'TRACKING_' + settings['date'], index_col=0)
    
    columns_track = df_track.columns
    columns_tracking = df_tracking.columns
    #columns_union = ['ProblemID', 'PatientID', 'Problem Summary', 'Problem']
    columns_union = columns_track


    if len(df_track)==0:
        ProblemIDMax=-1
        df_track = df_tracking[columns_union]
    else:
        ProblemIDMax = max([int(x) for x in list(df_track['ProblemID'])])
    
    ProblemIDInt = 0
    for index, row in df_tracking.iterrows():
        ProblemID = row['ProblemID']
        if not ProblemID == '':
            index = df_track['ProblemID'][df_track['ProblemID'] == ProblemID].index[0]
            for col in columns_union:
                df_track.loc[index,col] = row[col]
        else:
            ProblemIDInt = ProblemIDMax + 1
            ProblemIDMax = ProblemIDInt
            row['ProblemID'] = str(ProblemIDInt).zfill(6)
            row_new = pd.DataFrame('', index=[0], columns=columns_union)        
            for col in columns_union:
                row_new.loc[0,col] = row[col]
            df_track = df_track.append(row_new, ignore_index=True)
            df_tracking.loc[index,'ProblemID'] = str(ProblemIDInt).zfill(6)
    
    
    # Update master
    writer = pd.ExcelWriter(settings['filepath_master'], engine="openpyxl", mode="a")
    # Remove sheet if already exist
    sheet_name = 'TRACKING' + '_' + settings['date']
    workbook  = writer.book
    sheetnames = workbook.sheetnames
    if sheet_name in sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
        
    # Add patient ro master
    df_track.to_excel(writer, sheet_name=sheet_name)
    writer.save()
    
    # Update tracking
    writer = pd.ExcelWriter(settings['filepath_tracking'], engine="openpyxl", mode="a")
    # Remove sheet if already exist
    sheet_name = 'tracking table'
    workbook  = writer.book
    sheetnames = workbook.sheetnames
    if sheet_name in sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
        
    # Add patient to master
    df_tracking.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()


def orderMasterData(df_master, settings):
    """ Order columns of the master
        
    :param settings: Dictionary of settings
    :type settings: dict
    """
    # Reoder columns
    cols = df_master.columns.tolist()
    cols_new = settings['columns_first'] + [x for x in cols if x not in settings['columns_first']]
    df_master = df_master[cols_new]
    df_master = df_master.sort_values(['PatientID', 'StudyInstanceUID', 'SeriesInstanceUID'], ascending = (True, True, True))
    df_master.reset_index(inplace=True, drop=True)
    return df_master
    
def mergeMaster(settings):
    """ Merge master file
        
    :param settings: Dictionary of settings
    :type settings: dict
    """   
    
    print('Create master')
    # Read tables
    print('Read discharge_data')
    df_data = pd.read_pickle(settings['filepath_data'])
    print('Read discharge_pred')
    df_pred = pd.read_pickle(settings['filepath_prediction'])
    df_pred['CTA'] = df_pred['CTA'].astype('bool')
    print('Read discharge_reco')
    df_reco_load = pd.read_excel(settings['filepath_reco'], index_col=0)
    df_reco = pd.DataFrame()
    df_reco['RECO'] = df_reco_load['PredClass']
    df_reco['RECO_PROP'] = df_reco_load['Prop']
    print('Read discharge_rfc')
    df_rfc = pd.read_pickle(settings['filepath_rfc'])
    print('Read discharge_manual')
    df_manual = pd.read_pickle(settings['filepath_manual'])
    print('Read discharge_track')
    print('Create discharge_master')
    df_master = pd.concat([df_data, df_pred, df_rfc, df_manual, df_reco], axis=1)
    #df_master = pd.concat([df_data, df_pred, df_rfc, df_manual], axis=1)
    writer = pd.ExcelWriter(settings['filepath_master'], engine="openpyxl", mode="w")
    df_master.to_excel(writer, sheet_name = 'MASTER' + '_' + settings['date'])
    # Add patient data
    writer.save()


# def createMasterProcess(folderpath_master):
#     # Create master_process
#     date = folderpath_master.split('_')[-1]
#     filepath_master = os.path.join(folderpath_master, 'discharge_master_' + date + '.xlsx')
#     folderpath, filename, file_extension = splitFilePath(filepath_master)
#     filepath_process = os.path.join(folderpath, filename + '_process' + file_extension)
#     copyfile(filepath_master, filepath_process)


def formatMaster(settings):
    """ Format master table
        
    :param settings: Dictionary of settings
    :type settings: dict
    """
    print('Format master')
    format='ALL'
    # Read tables
    print('Read discharge_data')
    df_data = pd.read_pickle(settings['filepath_data'])
    print('Read discharge_pred')
    df_pred = pd.read_pickle(settings['filepath_prediction'])
    print('Read discharge_rfc')
    df_rfc = pd.read_pickle(settings['filepath_rfc'])
    print('Read discharge_manual')
    df_manual = pd.read_pickle(settings['filepath_manual'])
    print('Read discharge_track')
    #df_track = pd.read_excel(settings['filepath_master_track'], index_col=0)
    print('Read patient_data')
    df_patient = pd.read_pickle(settings['filepath_patient'])
    print('Create discharge_master')
    df_master = pd.read_excel(settings['filepath_master'], sheet_name='MASTER_' + settings['date'], index_col=0)
    df_master = orderMasterData(df_master, settings)

    # Save master
    writer = pd.ExcelWriter(settings['filepath_master'], engine="openpyxl", mode="a")
    sheet_name = 'MASTER_'+ settings['date']
    workbook  = writer.book
    sheetnames = workbook.sheetnames
    if sheet_name in sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
    df_master.to_excel(writer, sheet_name=sheet_name)
    writer.save()
    
    writer = pd.ExcelWriter(settings['filepath_master'], engine="openpyxl", mode="a")
    workbook  = writer.book
    
    colors=['A5A5A5', 'FFFF00', '70AD47', 'FFC000', '5B95F9', 'F86EEE']
    sheetnames = workbook.sheetnames
    
    for sheetname in sheetnames:
        sheet = workbook[sheetname]
        if 'DATA' in sheetname and (format=='ALL' or format=='DATA'):
            # Clear existing conditional_formatting list
            sheet.conditional_formatting = ConditionalFormattingList()
            sheet.data_validations.dataValidation = []
            # Highlight master
            columns = [x for x in range(0,df_data.shape[1])]
            sheet = highlight_columns(sheet, columns=columns , color=colors[0])
            # Freeze worksheet
            workbook[sheetname].freeze_panes = "D2"
            # Comment data
            sheet = setComment(sheet, columns=columns, comment='Please add a comment why the data has been changed!')
            # Add filter
            sheet.auto_filter.ref = sheet.dimensions
            # Draw border
            #sheet = setBorderFormula(sheet, formula='=$C1<>$C2', NumRows=df_data.shape[0], NumColumns=df_data.shape[1])
            
        if 'MASTER' in sheetname and (format=='ALL' or format=='MASTER'):

            # Clear existing conditional_formatting list
            sheet.conditional_formatting = ConditionalFormattingList()
            sheet.data_validations.dataValidation = []
            
            # Highlight master
            df_master_cols = list(df_master.columns)
            df_data_cols = list(df_data.columns)
            df_pred_cols = list(df_pred.columns)
            df_guide_cols = ['10-STEPS-GUIDE', '10-STEPS-GUIDE-COMMENT']
            df_rfc_cols = list(df_rfc.columns)
            df_manual_cols = list(df_manual.columns)
            #df_reco_cols = list(df_reco.columns)
            #df_track_cols = list(df_track.columns)
            sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in df_data_cols] , color=colors[0])
            sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in df_pred_cols] , color=colors[1])
            sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in df_rfc_cols] , color=colors[2])
            #sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in (df_rfc_cols + df_reco_cols)] , color=colors[2])
            #sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in df_reco_cols] , color=colors[2])
            
            sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in df_manual_cols] , color=colors[3])
            if df_guide_cols[0] in df_master_cols:
                sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in df_guide_cols] , color=colors[4])


            #sheet = highlight_columns(sheet, columns=[df_master_cols.index(col) for col in df_track_cols] , color=colors[4])
            
            # Comment  and access rights
            sheet = setComment(sheet, columns=[df_master_cols.index(col) for col in df_data_cols], comment='Please add a comment why the data has been changed!')
            sheet = setComment(sheet, columns=[df_master_cols.index(col) for col in df_pred_cols], comment='Do not change this data!')
            sheet = setAccessRights(sheet, columns=[df_master_cols.index(col) for col in ['RFCLabel']], promt='RFLabel!', promptTitle='DISCHARGE Scan Classes', formula1=scanClassesStr)
            sheet = setComment(sheet, columns=[df_master_cols.index(col) for col in ['RFCClass', 'RFCConfidence']], comment='Do not change this data!')
            sheet = setAccessRights(sheet, columns=[df_master_cols.index(col) for col in ['ClassManualCorrection']], promt='RFLabel!', promptTitle='DISCHARGE Scan Classes', formula1=scanClassesManualStr)
            sheet = setComment(sheet, columns=[df_master_cols.index(col) for col in ['Comment']], comment='Thank you for adding a comment!')
            sheet = setAccessRights(sheet, columns=[df_master_cols.index(col) for col in ['Image Quality']], promt='Image Quality', promptTitle='Image Quality', formula1=imageQualityStr)
            
            
            # Highlight based on modality
            setColorFormula(sheet, formula='$G2="CACS"', color="EE1111", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            setColorFormula(sheet, formula='$G2="CTA"', color="00B050", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            setColorFormula(sheet, formula='$G2="NCS_CACS"', color="0070C0", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            setColorFormula(sheet, formula='$G2="NCS_CTA"', color="FFC000", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            setColorFormula(sheet, formula='$G2="UNDEFINED"', color="FF33CC", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            setColorFormula(sheet, formula='$G2="OTHER"', color="FC74F2", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            
            
            #setColorFormula(sheet, formula='ISBLANK($CC1)', color="000000", colorrange="A1:AI45000")
            
            # Highligt based on confidence score
            # setColorFormula(sheet, formula='$I1<0.5', color="EE1111", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            # setColorFormula(sheet, formula='AND($I1<0.9, $I1>0.3)', color="FFFF00", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            # setColorFormula(sheet, formula='$I1>0.9', color="00B050", NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            
            # Freeze worksheet
            workbook[sheetname].freeze_panes = "D2"
            # Draw border
            sheet = setBorderFormula(sheet, formula='=$C1<>$C2', NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            #sys.exit()
            # Add filter
            sheet.auto_filter.ref = sheet.dimensions

        if 'PATIENT' in sheetname and (format=='ALL' or format=='PATIENT') and (not 'PATIENT_STATUS_CONF' in sheetname):
            # Clear existing conditional_formatting list
            sheet.conditional_formatting = ConditionalFormattingList()
            sheet.data_validations.dataValidation = []
            df_patient_cols = list(df_patient.columns)
            sheet = highlight_columns(sheet, columns=[df_patient_cols.index(col) for col in df_patient_cols] , color=colors[4])
            #patient_status_manual = ['OK', 'EXCLUDED', 'UNDEFINED']
            sheet = setComment(sheet, columns=[df_patient_cols.index(col) for col in df_patient_cols[0:-2]], comment='Do not change this data!')
            sheet = setAccessRights(sheet, columns=[df_patient_cols.index(col) for col in ['STATUS_MANUAL_CORRECTION']], promt='PatientLabel', promptTitle='DISCHARGE Patient label', formula1=patient_status_manualStr)
            sheet = setComment(sheet, columns=[df_patient_cols.index(col) for col in ['COMMENT']], comment='Thank you for adding a comment!')
            # Freeze worksheet
            workbook[sheetname].freeze_panes = "D2"
            # Draw border
            #sheet = setBorderFormula(sheet, formula='=$C1<>$C2', NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            # Add filter
            sheet.auto_filter.ref = sheet.dimensions

        if 'PATIENT_STATUS_CONF' in sheetname and (format=='ALL' or format=='PATIENT'):
            # Clear existing conditional_formatting list
            sheet.conditional_formatting = ConditionalFormattingList()
            sheet.data_validations.dataValidation = []
            df_patient_cols = list(df_patient.columns)
            sheet = highlight_columns(sheet, columns=[df_patient_cols.index(col) for col in df_patient_cols] , color=colors[4])
            #patient_status_manual = ['OK', 'EXCLUDED', 'UNDEFINED']
            sheet = setComment(sheet, columns=[df_patient_cols.index(col) for col in df_patient_cols[0:-2]], comment='Do not change this data!')
            sheet = setAccessRights(sheet, columns=[df_patient_cols.index(col) for col in ['STATUS_MANUAL_CORRECTION']], promt='PatientLabel', promptTitle='DISCHARGE Patient label', formula1=patient_status_manualStr)
            sheet = setComment(sheet, columns=[df_patient_cols.index(col) for col in ['COMMENT']], comment='Thank you for adding a comment!')
            # Freeze worksheet
            workbook[sheetname].freeze_panes = "D2"
            # Draw border
            #sheet = setBorderFormula(sheet, formula='=$C1<>$C2', NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            # Add filter
            sheet.auto_filter.ref = sheet.dimensions
            
            
        if 'TRACKING' in sheetname and (format=='ALL' or format=='TRACKING'):
            
            dft = pd.read_excel(settings['filepath_master'], sheet_name=sheetname, index_col=0)
            
            # Clear existing conditional_formatting list
            sheet.conditional_formatting = ConditionalFormattingList()
            sheet.data_validations.dataValidation = []
            #df_track_cols = list(df_track.columns)
            df_track_cols = list(dft.columns)
            sheet = highlight_columns(sheet, columns=[df_track_cols.index(col) for col in df_track_cols] , color=colors[4])
            Problem_Summary = ['UNDEFINED',
                                'Big problem',
                                'Date of the CT images are wrong',
                                'Missing ICA Images',
                                'ICA Problem']
            

            Problem_SummaryStr = '"' + 'UNDEFINED,' + 'Big problem,' + 'Date of the CT images are wrong,' + 'Missing ICA Images,' + 'ICA Problem,' + '"'

            sheet = setAccessRights(sheet, columns=[df_track_cols.index(col) for col in ['Problem Summary']], promt='Problem!', promptTitle='DISCHARGE Problem Summary', formula1=Problem_SummaryStr)
            # Freeze worksheet
            workbook[sheetname].freeze_panes = "D2"
            # Draw border
            #sheet = setBorderFormula(sheet, formula='=$C1<>$C2', NumRows=df_master.shape[0], NumColumns=df_master.shape[1])
            # Add filter
            sheet.auto_filter.ref = sheet.dimensions

    writer.save()
    

def createStudy(settings):
    """ Create StudyInstance realted statistics about missing images
        
    :param settings: Dictionary of settings
    :type settings: dict
    """
    print('Create StudyInstanceID table.')
    
    conf=True

    def getConf(conf, NumSeries=1):
        conf = conf.sort_values(ascending=False)
        if len(conf)==0:
            return 0
        if len(conf) <= NumSeries:
            return conf.min()
        else:
            return conf[0:NumSeries].min()
        
    CACS_NUM_MIN = 1
    CACS_IR_NUM_MIN = 0
    CACS_FBP_NUM_MIN = 0
    
    CTA_NUM_MIN = 1
    CTA_IR_NUM_MIN = 0
    CTA_FBP_NUM_MIN = 0
    
    NCS_CACS_NUM_MIN = 1
    NCS_CACS_IR_NUM_MIN = 0
    NCS_CACS_FBP_NUM_MIN = 0
    
    NCS_CTA_NUM_MIN = 1
    NCS_CTA_IR_NUM_MIN = 0
    NCS_CTA_FBP_NUM_MIN = 0

    # date = folderpath_master.split('_')[-1]
    # folderpath_components = os.path.join(folderpath_master, 'discharge_components_' + date)
    # filepath_pred = os.path.join(folderpath_components, 'discharge_pred_' + date + '.xlsx')
    # filepath_master_data = os.path.join(folderpath_components, 'discharge_master_data_' + date + '.xlsx')
    # filepath_patient = os.path.join(folderpath_components, 'discharge_patient_' + date + '.xlsx')
    # filepath_patient_conf = os.path.join(folderpath_components, 'discharge_patient_conf_' + date + '.xlsx')
    #filepath_master = os.path.join(folderpath_master, 'discharge_master_' + date + '.xlsx')
    # if master_process==False:
    #     filepath_master = os.path.join(folderpath_master, 'discharge_master_' + date + '.xlsx')
    # else:
    #     filepath_master_tmp = os.path.join(folderpath_master, 'discharge_master_' + date + '.xlsx')
    #     folderpath, filename, file_extension = splitFilePath(filepath_master_tmp)
    #     filepath_master = os.path.join(folderpath, filename + '_process' + file_extension)
    
    df_master = pd.read_excel(settings['filepath_master'], sheet_name='MASTER_'+ settings['date'], index_col=0)
    df_master.sort_index(inplace=True)
    #df_master_data = pd.read_excel(filepath_master_data, index_col=0)
    df_PatientID = pd.DataFrame(columns=['Site', 'PatientID', 'StudyInstanceUID', 'Modality', 'AcquisitionDate',
                                       'CACS_NUM', 'CACS_FBP_NUM', 'CACS_IR_NUM', 
                                       'CTA_NUM', 'CTA_FBP_NUM', 'CTA_IR_NUM', 
                                       'NCS_CACS_NUM', 'NCS_CACS_FBP_NUM', 'NCS_CACS_IR_NUM', 
                                       'NCS_CTA_NUM', 'NCS_CTA_FBP_NUM', 'NCS_CTA_IR_NUM',
                                       'ITT', 'STATUS', 'STATUS_MANUAL_CORRECTION', 'COMMENT'])
    df_PatientID_conf = pd.DataFrame(columns=['Site', 'PatientID', 'StudyInstanceUID', 'Modality', 'AcquisitionDate',
                                       'CACS_CONF', 'CACS_FBP_CONF', 'CACS_IR_CONF', 
                                       'CTA_CONF', 'CTA_FBP_CONF', 'CTA_IR_CONF', 
                                       'NCS_CACS_CONF', 'NCS_CACS_FBP_CONF', 'NCS_CACS_IR_CONF', 
                                       'NCS_CTA_CONF', 'NCS_CTA_FBP_CONF', 'NCS_CTA_IR_CONF',
                                       'ITT', 'STATUS', 'STATUS_MANUAL_CORRECTION', 'COMMENT', 'Conf<0.5'])

    # Filter study list
    func = lambda x: datetime.strptime(x, '%Y%m%d')
    patients = df_master['PatientID'].unique()
    firstdateList = []
    study_list = []

    for patient in patients:
        #if patient == '02-INN-0020':
        #    sys.exit()
        df_patient = df_master[(df_master['PatientID']==patient) &  (df_master['Modality']=='CT')]
        if len(df_patient)>0:
            #sys.exit()
            firstdate = df_patient['1. Date of CT scan'].iloc[0]
            
            studydate = df_patient['StudyDate']
            # Convert string to date and replace in df_patient
            studydate = studydate.apply(lambda x: datetime.datetime.strptime(str(x), '%Y%m%d'))
            df_patient['StudyDate'] = studydate
            
            if not pd.isnull(firstdate):
                df_study = df_patient[studydate == firstdate]
                if len(df_study)>0:
                    #df_study_id = df_study['StudyInstanceUID'].iloc[0]
                    df_study_id_list = df_study['StudyInstanceUID']
                else:
                    print('PROBLEM: Patient ' + patient + ' "1. Date of CT scan" not consistent with "StudyDate"')
                    df_study = df_patient
                    df_study = df_study.sort_values(by='StudyDate')
                    df_study_id_list = df_study['StudyInstanceUID'][0:1]
            else:
                print('Patient: ' + patient + ' does not have a 1. Date of CT scan')
                df_study = df_patient
                df_study = df_study.sort_values(by='StudyDate')
                df_study_id = df_study['StudyInstanceUID'].iloc[0]
                df_study_id_list = df_study['StudyInstanceUID'][0:1]
            study_list.append(df_study_id_list)
                  

    for studyIDList in study_list:
        # if '1.2.40.0.13.1.270104688921897647052181046394362664786' in list(studyIDList):
        #     sys.exit()
        #data = df_master_data[df_master_data['PatientID']==patientID]
        #if data['Modality'].iloc[0]=='CT':
        df_study = df_master[df_master['StudyInstanceUID'].isin(studyIDList)]
        # Extract CACS_NUM information
        CACS_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='CACS')
        #CACS_AUTO_OK = (df_study['CACS']) & (df_study['ITT']<2)
        CACS_AUTO_OK = (df_study['RFCClass']=='CACS') & (df_study['ITT']<2)
        CACS_NUM = (CACS_MANUAL_OK & CACS_AUTO_OK).sum()
        CACS_CONF = getConf(df_study[CACS_MANUAL_OK & CACS_AUTO_OK]['RFCConfidence'], NumSeries=CACS_NUM_MIN)
        
        # Extract CACS_FBP_NUM information
        CACS_FBP_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='CACS')
        #CACS_FBP_AUTO_OK = (df_study['CACS'])  & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        CACS_FBP_AUTO_OK = (df_study['RFCClass']=='CACS')  & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        CACS_FBP_NUM = (CACS_FBP_MANUAL_OK & CACS_FBP_AUTO_OK).sum()
        CACS_FBP_CONF = getConf(df_study[CACS_MANUAL_OK & CACS_FBP_AUTO_OK]['RFCConfidence'], NumSeries=CACS_FBP_NUM_MIN)
        # Extract CACS_IR_NUM information
        CACS_IR_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='CACS')
        #CACS_IR_AUTO_OK = (df_study['CACS']) & (df_study['RECO']=='IR') & (df_study['ITT']<2)
        CACS_IR_AUTO_OK = (df_study['RFCClass']=='CACS') & (df_study['RECO']=='IR') & (df_study['ITT']<2)
        CACS_IR_NUM = (CACS_IR_MANUAL_OK & CACS_IR_AUTO_OK).sum()
        CACS_IR_CONF = getConf(df_study[CACS_MANUAL_OK & CACS_IR_AUTO_OK]['RFCConfidence'], NumSeries=CACS_IR_NUM_MIN)
        # Extract CTA_NUM information
        CTA_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='CTA')
        #CTA_AUTO_OK = (df_study['CTA']) & (df_study['ITT']<2)
        CTA_AUTO_OK = (df_study['RFCClass']=='CTA') & (df_study['ITT']<2)
        CTA_NUM = (CTA_MANUAL_OK & CTA_AUTO_OK).sum()  
        CTA_CONF = getConf(df_study[CTA_MANUAL_OK & CTA_AUTO_OK]['RFCConfidence'], NumSeries=CTA_NUM_MIN)
        # Extract CTA_FBP_NUM information
        CTA_FBP_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='CTA')
        #CTA_FBP_AUTO_OK = (df_study['CTA'])  & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        CTA_FBP_AUTO_OK = (df_study['RFCClass']=='CTA')  & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        CTA_FBP_NUM = (CTA_FBP_MANUAL_OK & CTA_FBP_AUTO_OK).sum()  
        CTA_FBP_CONF = getConf(df_study[CTA_FBP_MANUAL_OK & CTA_FBP_AUTO_OK]['RFCConfidence'], NumSeries=CTA_FBP_NUM_MIN)
        # Extract CTA_IR_NUM information
        CTA_IR_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='CTA')
        #CTA_IR_AUTO_OK = (df_study['CTA'])  & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        CTA_IR_AUTO_OK = (df_study['RFCClass']=='CTA')  & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        CTA_IR_NUM = (CTA_IR_MANUAL_OK & CTA_IR_AUTO_OK).sum()          
        CTA_IR_CONF = getConf(df_study[CTA_IR_MANUAL_OK & CTA_IR_AUTO_OK]['RFCConfidence'], NumSeries=CTA_IR_NUM_MIN)
        # Extract NCS_CACS_NUM information
        NCS_CACS_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='NCS_CACS')
        #NCS_CACS_AUTO_OK = (df_study['NCS_CACS']) & (df_study['ITT']<2)
        NCS_CACS_AUTO_OK = (df_study['RFCClass']=='NCS_CACS') & (df_study['ITT']<2)
        NCS_CACS_NUM = (NCS_CACS_MANUAL_OK & NCS_CACS_AUTO_OK).sum()    
        NCS_CACS_CONF = getConf(df_study[NCS_CACS_MANUAL_OK & NCS_CACS_AUTO_OK]['RFCConfidence'], NumSeries=NCS_CACS_NUM_MIN)
        # Extract NCS_CACS_FBP_NUM information
        NCS_CACS_FBP_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='NCS_CACS')
        #NCS_CACS_FBP_AUTO_OK = (df_study['NCS_CACS']) & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        NCS_CACS_FBP_AUTO_OK = (df_study['RFCClass']=='NCS_CACS') & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        NCS_CACS_FBP_NUM = (NCS_CACS_FBP_MANUAL_OK & NCS_CACS_FBP_AUTO_OK).sum()  
        NCS_CACS_FBP_CONF = getConf(df_study[NCS_CACS_FBP_MANUAL_OK & NCS_CACS_FBP_AUTO_OK]['RFCConfidence'], NumSeries=NCS_CACS_FBP_NUM_MIN)
        # Extract NCS_CACS_IR_NUM information
        NCS_CACS_IR_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='NCS_CACS')
        #NCS_CACS_IR_AUTO_OK = (df_study['NCS_CACS']) & (df_study['RECO']=='IR') & (df_study['ITT']<2)
        NCS_CACS_IR_AUTO_OK = (df_study['RFCClass']=='NCS_CACS') & (df_study['RECO']=='IR') & (df_study['ITT']<2)
        NCS_CACS_IR_NUM = (NCS_CACS_IR_MANUAL_OK & NCS_CACS_IR_AUTO_OK).sum()   
        NCS_CACS_IR_CONF = getConf(df_study[NCS_CACS_IR_MANUAL_OK & NCS_CACS_IR_AUTO_OK]['RFCConfidence'], NumSeries=NCS_CACS_IR_NUM_MIN)
        # Extract NCS_CACS_NUM information
        NCS_CTA_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='NCS_CTA')
        #NCS_CTA_AUTO_OK = (df_study['NCS_CTA']) & (df_study['ITT']<2)
        NCS_CTA_AUTO_OK = (df_study['RFCClass']=='NCS_CTA') & (df_study['ITT']<2)
        NCS_CTA_NUM = (NCS_CTA_MANUAL_OK & NCS_CTA_AUTO_OK).sum() 
        NCS_CTA_CONF = getConf(df_study[NCS_CTA_MANUAL_OK & NCS_CTA_AUTO_OK]['RFCConfidence'], NumSeries=NCS_CTA_NUM_MIN)
        # Extract NCS_CTA_FBP_NUM information
        NCS_CTA_FBP_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='NCS_CTA')
        #NCS_CTA_FBP_AUTO_OK = (df_study['NCS_CTA']) & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        NCS_CTA_FBP_AUTO_OK = (df_study['RFCClass']=='NCS_CTA') & (df_study['RECO']=='FBP') & (df_study['ITT']<2)
        NCS_CTA_FBP_NUM = (NCS_CTA_FBP_MANUAL_OK & NCS_CTA_FBP_AUTO_OK).sum() 
        NCS_CTA_FBP_CONF = getConf(df_study[NCS_CTA_FBP_MANUAL_OK & NCS_CTA_FBP_AUTO_OK]['RFCConfidence'], NumSeries=NCS_CTA_FBP_NUM_MIN)
        # Extract NCS_CTA_IR_NUM information
        NCS_CTA_IR_MANUAL_OK = (df_study['ClassManualCorrection']=='UNDEFINED') | (df_study['ClassManualCorrection']=='NCS_CTA')
        #NCS_CTA_IR_AUTO_OK = (df_study['NCS_CTA']) & (df_study['RECO']=='IR') & (df_study['ITT']<2)
        NCS_CTA_IR_AUTO_OK = (df_study['RFCClass']=='NCS_CTA') & (df_study['RECO']=='IR') & (df_study['ITT']<2)
        NCS_CTA_IR_NUM = (NCS_CTA_IR_MANUAL_OK & NCS_CTA_IR_AUTO_OK).sum() 
        NCS_CTA_IR_CONF = getConf(df_study[NCS_CTA_IR_MANUAL_OK & NCS_CTA_IR_AUTO_OK]['RFCConfidence'], NumSeries=NCS_CTA_IR_NUM_MIN)
        
        PATIENTID = df_study['PatientID'].iloc[0]
        SITE = df_study['Site'].iloc[0]
        ITT = df_study['ITT'].iloc[0]
        modality = df_study['Modality'].iloc[0]
        DATE = df_study['AcquisitionDate'].iloc[0]
        STUDYID = ','.join(studyIDList)
        #AcquisitionDate = df_study['AcquisitionDate'].iloc[0]
        
        # Check patient scenario
        CACS_OK = CACS_NUM>=CACS_NUM_MIN
        CTA_OK = CTA_NUM>=CTA_NUM_MIN
        NCS_CACS_OK = NCS_CACS_NUM>=NCS_CACS_NUM_MIN
        NCS_CTA_OK = NCS_CTA_NUM>=NCS_CTA_NUM_MIN
        
        if CACS_OK and CTA_OK and NCS_CACS_OK and NCS_CTA_OK:
            status = 'OK'
        elif ITT==2:
            status = 'EXCLUDED'
        elif not modality=='CT':
            status = 'NOT CT MODALITY'
        else:
            status=''
            if not CACS_OK:
                status = status + 'MISSING_CACS, '
            if not CTA_OK:
                status = status + 'MISSING_CTA, '
            if not NCS_CACS_OK:
                status = status + 'MISSING_NCS_CACS, '
            if not NCS_CTA_OK:
                status = status + 'MISSING_NCS_CTA, '
            
        STATUS_MANUAL_CORRECTION = 'UNDEFINED'
        COMMENT = ''
        if CACS_CONF<0.5 or CTA_CONF<0.5 or NCS_CACS_CONF<0.5 or NCS_CTA_CONF<0.5:
            ConfSmall = 'check the series'
        else:
            ConfSmall = 'confidence is high'
            
        df_PatientID = df_PatientID.append({'Site': SITE, 'PatientID': PATIENTID, 'StudyInstanceUID': STUDYID, 'Modality': modality, 'AcquisitionDate': DATE, 'CACS_NUM': CACS_NUM, 'CACS_FBP_NUM': CACS_FBP_NUM, 'CACS_IR_NUM': CACS_IR_NUM,
                           'CTA_NUM': CTA_NUM, 'CTA_FBP_NUM': CTA_FBP_NUM, 'CTA_IR_NUM': CTA_IR_NUM,
                           'NCS_CACS_NUM': NCS_CACS_NUM, 'NCS_CACS_FBP_NUM': NCS_CACS_FBP_NUM, 'NCS_CACS_IR_NUM': NCS_CACS_IR_NUM,
                           'NCS_CTA_NUM': NCS_CTA_NUM, 'NCS_CTA_FBP_NUM': NCS_CTA_FBP_NUM, 'NCS_CTA_IR_NUM': NCS_CTA_IR_NUM,
                           'ITT': ITT, 'STATUS': status, 'STATUS_MANUAL_CORRECTION': STATUS_MANUAL_CORRECTION, 'COMMENT': COMMENT}, ignore_index=True)
        df_PatientID_conf = df_PatientID_conf.append({'Site': SITE, 'PatientID': PATIENTID, 'StudyInstanceUID': STUDYID, 'Modality': modality, 'AcquisitionDate': DATE, 'CACS_CONF': CACS_CONF, 'CACS_FBP_CONF': CACS_FBP_CONF, 'CACS_IR_CONF': CACS_IR_CONF,
                           'CTA_CONF': CTA_CONF, 'CTA_FBP_CONF': CTA_FBP_CONF, 'CTA_IR_CONF': CTA_IR_CONF,
                           'NCS_CACS_CONF': NCS_CACS_CONF, 'NCS_CACS_FBP_CONF': NCS_CACS_FBP_CONF, 'NCS_CACS_IR_CONF': NCS_CACS_IR_CONF,
                           'NCS_CTA_CONF': NCS_CTA_CONF, 'NCS_CTA_FBP_CONF': NCS_CTA_FBP_CONF, 'NCS_CTA_IR_CONF': NCS_CTA_IR_CONF,
                           'ITT': ITT, 'STATUS': status, 'STATUS_MANUAL_CORRECTION': STATUS_MANUAL_CORRECTION, 'COMMENT': COMMENT, 'Conf<0.5': ConfSmall}, ignore_index=True)
        
    df_PatientID.to_pickle(settings['filepath_patient'])
    df_PatientID_conf.to_pickle(settings['filepath_patient_conf'])

    writer = pd.ExcelWriter(settings['filepath_master'], engine="openpyxl", mode="a")
    # Remove sheet if already exist
    sheet_name = 'PATIENT_STATUS_' + settings['date']
    workbook  = writer.book
    sheetnames = workbook.sheetnames
    if sheet_name in sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
    df_PatientID.to_excel(writer, sheet_name=sheet_name)
        
    if conf:
        sheet_name = 'PATIENT_STATUS_CONF_' + settings['date']
        if sheet_name in sheetnames:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
        df_PatientID_conf.to_excel(writer, sheet_name=sheet_name)
        
    # Add patient to master
    writer.save()
    
    
def extractHist(settings):
    """ Extract histogram of images
        
    :param settings: Dictionary of settings
    :type settings: dict
    """
    
    df = pd.read_excel(settings['filepath_master'], sheet_name ='MASTER_' + settings['date'], index_col=0)
    df.sort_index(inplace=True)
    bins = 100   
    columns =['SeriesInstanceUID', 'Count', 'CLASS'] + [str(x) for x in range(0,bins)]
    if os.path.exists(settings['filepath_hist']):
        dfHist = pd.read_pickle(settings['filepath_hist'])
    else:
        dfHist = pd.DataFrame('', index=np.arange(len(df)), columns=columns)
    
    start = 0
    end = len(df)
    #end = 1200
    print('Press "ctrl + e" to stop execution.')
    pbar = tqdm(total=len(df))
    pbar.set_description("Extract histograms")
    for index, row in df[start:end].iterrows():
        #print('index', index)   
        pbar.update(1)
        if dfHist.iloc[index,0]=='':
            if keyboard.is_pressed('ctrl+e'):
                print('Button "ctrl + e" pressed to exit execution.')
                sys.exit()
            StudyInstanceUID=row['StudyInstanceUID']
            PatientID=row['PatientID']
            SeriesInstanceUID=row['SeriesInstanceUID']    
            if not (row['Modality']=='CT' or row['Modality']=='OT'):
                print('Series modality is not CT')
                dfHist.loc[index,'SeriesInstanceUID'] = SeriesInstanceUID
                dfHist.loc[index,'Count'] = -1
                dfHist.loc[index,'CLASS'] = row['CLASS']
                dfHist.iloc[index,3:] = np.ones((1, bins))*-1
            else:
                try:
                    patient=CTPatient(StudyInstanceUID, PatientID)
                    series = patient.loadSeries(settings['folderpath_discharge'], SeriesInstanceUID, None)
                    image = series.image.image()
                    hist = np.histogram(image, bins=bins, range=(-2500, 3000))[0]
                    dfHist.loc[index,'SeriesInstanceUID'] = SeriesInstanceUID
                    dfHist.loc[index,'Count'] = image.shape[0]
                    dfHist.loc[index,'CLASS'] = row['CLASS']
                    dfHist.iloc[index,3:] = hist
                except:
                    print('Error index', index)
                    dfHist.loc[index,'SeriesInstanceUID'] = SeriesInstanceUID
                    dfHist.loc[index,'Count'] = -1
                    dfHist.loc[index,'CLASS'] = row['CLASS']
                    dfHist.iloc[index,3:] = np.ones((1, bins))*-1
        if index % 10 == 0:
            dfHist.to_pickle(settings['filepath_hist'])
    pbar.close()
    dfHist.to_pickle(settings['filepath_hist'])

def mergeManualSelection(settings):
    """ Merge manual selected series
        
    :param settings: Dictionary of settings
    :type settings: dict
    """
    
    filepath_master = settings['filepath_master']
    discharge_manual_selection = settings['folderpath_manual_selection']
    df_master = pd.read_excel(filepath_master, sheet_name='MASTER_'+ settings['date'], index_col=0)
    
    # Extract manual files
    centers = defaultdict(lambda:None, {})
    filepath_manual = glob(discharge_manual_selection + '/*.xlsx')
    
    # Check PALL file
    for file in filepath_manual:
        if 'PALL' in file:
            for c in settings['centers']:
                centers[c] = file
    
    if len(centers)==0:
        for file in filepath_manual:
            filesplit = file[0:-5].split("_")
            for st in filesplit:
                if st[0]=='P' and len(st)==3:
                    centers[st] = file

    
    # Update master for each center
    for center in centers.keys():
        
        #if center=='P01' or center=='P02':
        print('Processing center', center)
        filepath_manual = centers[center]
        columns_copy=['SeriesInstanceUID', 'ClassManualCorrection', 'Comment', 'Responsible Person', 'RFCLabel', 'RFCClass']
        sheet_name = 'MASTER_'+ settings['date']
        df_manual = pd.read_excel(filepath_manual, sheet_name=sheet_name, index_col=0)
        df_manual.replace(to_replace=[np.nan], value='', inplace=True)
        df_manual_P = df_manual[df_manual['Site']==center]
        df_manual_P.replace(to_replace=['CACSExtended'], value='CACS', inplace=True)
        df_manual_P.replace(to_replace=['CTAExtended'], value='CTA', inplace=True)
        df_manual_P.replace(to_replace=['NCS_CACSExtended'], value='NCS_CACS', inplace=True)
        df_manual_P.replace(to_replace=['NCS_CTAExtended'], value='NCS_CTA', inplace=True)
        df_manual_P = df_manual_P[columns_copy]
        
        df_merge = df_master.merge(df_manual_P, on=['SeriesInstanceUID', 'SeriesInstanceUID'], how='left')
        
        idx_master = df_master['Site']==center
        df_master['RFCClass'][idx_master] = df_merge['RFCClass_y'][idx_master]
        df_master['RFCLabel'][idx_master] = df_merge['RFCClass_y'][idx_master]
        df_master['RFCLabel'][idx_master] = df_merge['RFCClass_y'][idx_master]
        df_master['Image Quality'][idx_master] = df_merge['Image Quality'][idx_master]
        df_master['Responsible Person'][idx_master] = df_merge['Responsible Person_y'][idx_master]
        df_master['Comment'][idx_master] = df_merge['Comment_y'][idx_master]
        df_master['ClassManualCorrection'][idx_master] = df_merge['ClassManualCorrection_y'][idx_master]



    idx = ~(df_master['ClassManualCorrection']=='UNDEFINED')
    df_master['RFCClass'][idx] = df_master['ClassManualCorrection']
    df_master['RFCLabel'][idx] = df_master['ClassManualCorrection']
        
    # Save master
    writer = pd.ExcelWriter(filepath_master, engine="openpyxl", mode="a")
    sheet_name = 'MASTER_'+ settings['date']
    workbook  = writer.book
    sheetnames = workbook.sheetnames
    if sheet_name in sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
    df_master.to_excel(writer, sheet_name=sheet_name)
    writer.save()

def checkAutomaticManual(settings):
    """ Check manual set boundaries automatically
        
    :param settings: Dictionary of settings
    :type settings: dict
    """
    
    filepath_master = settings['filepath_master']
    df_master = pd.read_excel(filepath_master, sheet_name='MASTER_'+ settings['date'], index_col=0)
    df_master_copy = df_master.copy()
    df_master_copy.replace(to_replace=[np.nan], value=0.0, inplace=True)
    
    # Check Multi-Slice CTs
    idx_multi = (df_master_copy['NumberOfFrames'] > 1) & (df_master_copy['Count'] > 1)
    df_master['ClassManualCorrection'][idx_multi] = 'PROBLEM'
    df_master['Comment'][idx_multi] = 'Multiple Multi-Slice-CTs under one SeriesInstanceUID.'
    df_master['Responsible Person'][idx_multi] = 'BF_AUT'
    
    # Check CTs with number of slices smaller 15
    idx_count = (df_master_copy['NumberOfFrames'] == 0.0) & (df_master_copy['Count'] < 15)
    df_master['ClassManualCorrection'][idx_count] = 'OTHER'
    df_master['Comment'][idx_count] = 'Selected as other becaus number of slices is smaller 15'
    df_master['Responsible Person'][idx_count] = 'BF_AUT'
            
    # Save master
    writer = pd.ExcelWriter(filepath_master, engine="openpyxl", mode="a")
    sheet_name = 'MASTER_'+ settings['date']
    workbook  = writer.book
    sheetnames = workbook.sheetnames
    if sheet_name in sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
    df_master.to_excel(writer, sheet_name=sheet_name)
    writer.save()
    
    
def createMaster():
    """ Create master file
    """
    
    # Load settings
    filepath_settings = 'H:/cloud/cloud_data/Projects/DISCHARGEMaster/data/settings.json'
    settings=initSettings()
    saveSettings(settings, filepath_settings)
    settings = fillSettingsTags(loadSettings(filepath_settings))
    
    # ExtracextractHistt histograms
    extractHist(settings)
    # Extract dicom tags
    #extractDICOMTags(settings, NumSamples=None)
    # Create tables
    checkTables(settings)
    # Create data
    createData(settings, NumSamples=None)
    # Create random forest classification columns
    createRFClassification(settings)
    # Create manual selection
    createManualSelection(settings)
    # Create prediction 
    createPredictions(settings)
    # Merge master 
    mergeMaster(settings)
    # Create tracking table 
    createTrackingTable(settings)
    # Init RF classifier
    initRFClassification(settings)
    # Update manual selection
    mergeManualSelection(settings)
    #Classifie based on manual selection
    classifieRFClassification(settings)
    # Check manual selection
    checkAutomaticManual(settings)
    # Filter according to 10StepsGuide
    filer10StepsGuide(settings)
    # Merge study sheet 
    createStudy(settings)
    # Format master
    formatMaster(settings)
    
    #df_reco_load = pd.read_excel('H:/cloud/cloud_data/Projects/DISCHARGEMaster/data/discharge_master/discharge_master_01092020/discharge_components_01092020/discharge_reco_pred_01092020.xlsx')
    
    # settings['filepath_sn'] = 'H:/cloud/cloud_data/Projects/DISCHARGEMaster/src/reco/sn.pkl'
    # filt = RecoFilter()
    # filt.extractSignalToNoise(settings, NumSamples=[0,1000])
    
    
    # clf, confidence, C, ACC, pred = filt.trainNoiseTree(settings)
    
    # dfSN = pd.read_pickle(settings['filepath_sn'])
    
    
    
    
    # df_master = pd.read_excel(settings['filepath_master'], sheet_name = 'MASTER' + '_' + settings['date'], index_col=0)
    # df_cacs = df_master[df_master['RFCClass']=='CACS']
    
    # patient=''
    # numlist=[]
    # for index, row in df_cacs.iterrows():
    #     if p==row['PatientID']:
    #         num=num+1
    #     else:
    #         numlist.append(num)
    #         num=1
    #         p=row['PatientID']
    
    
    # PatientID = '01-BER-0012'
    # StudyInstanceUID = '1.2.840.113619.6.95.31.0.3.4.1.1018.13.10329788'
    # SeriesInstanceUID = '1.2.392.200036.9116.2.6.1.37.2426555318.1461798187.314816'
    # patient = CTPatient(StudyInstanceUID, PatientID)
    # series2 = patient.loadSeries(settings['folderpath_discharge'], SeriesInstanceUID, None)
    # image0 = series2.image
    # s0 = image0.image()[25,:,:]
    # r0=s0[200:250,250:300]
    # m0=np.mean(r0)  
    # s0=np.std(r0)
    # sn0=m0/s0
    
    # feature0 = self.sgnalToNoise(image0) 
    
    
    
    # PatientID = '01-BER-0014'
    # StudyInstanceUID = '1.2.840.113619.6.95.31.0.3.4.1.1018.13.10347678'
    # SeriesInstanceUID = '1.2.392.200036.9116.2.6.1.37.2426555318.1462922209.252410'
    # patient = CTPatient(StudyInstanceUID, PatientID)
    # series1 = patient.loadSeries(settings['folderpath_discharge'], SeriesInstanceUID, None)
    # image1 = series1.image
    # s1 = image1.image()[25,:,:]
    # r1=s1[200:250,250:300]
    # m1=np.mean(r1)  
    # s1=np.std(r1)
    # sn1=m1/s1
    
    # feature1 = self.sgnalToNoise(image1) 
    
    # Load data
    df_hist = pd.read_pickle('H:/cloud/cloud_data/Projects/DISCHARGEMaster/data/discharge_master/discharge_master_01092020/discharge_sources_01092020/discharge_hist_01092020.pkl')
    columns = [str(i) for i in range(100)]
    
    # Filter  empty arrays
    df_col = df_hist[columns]
    idx_empty = df_col['0']>-1
    df_filt0 = df_col[idx_empty]
    df_hist0 = df_hist[idx_empty]
    
    # Filter dublicates
    idx_dub = df_filt0.duplicated(keep=False)
    df_hist1 = df_hist0[idx_dub]
    
    # Extract dublicates
    pbar = tqdm(total=len(df_hist1))
    pbar.set_description("Extract dublicates")
    df_dublicates = pd.DataFrame(columns=['SeriesInstanceUID', 'dublicates'])
    for index0, row0 in df_hist1.iterrows():
        pbar.update(1)
        c0 = list(row0[columns])
        dub_list=[]
        for index1, row1 in df_hist1.iterrows():
            if row0['SeriesInstanceUID']!=row1['SeriesInstanceUID']:
                c1 = list(row1[columns])
                if c0==c1:
                    dub_list.append(row1['SeriesInstanceUID'])
        if len(dub_list)>0:
            df_dublicates = df_dublicates.append(dict({'SeriesInstanceUID': row0['SeriesInstanceUID'], 'dublicates': dub_list}), ignore_index=True)
    pbar.close()
    dfHist.to_pickle(settings['filepath_hist'])
                
                
                
                
                
    #         c1 = list(row1[columns])
    #         if row['SeriesInstanceUID']=='1.2.40.0.13.1.109250245863884995619126169614607621221':
    #             row0 = row
    #         if row['SeriesInstanceUID']=='1.2.40.0.13.1.53338717653213650310124647000719089661':
    #             row1 = row
            
            
            
            
            
            