# -*- coding: utf-8 -*-
"""
Created on Wed May 13 13:59:31 2020

@author: bernifoellmer
"""

import sys, os
import pandas as pd
import openpyxl
import ntpath
import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Color, Border, Side
from openpyxl.styles import colors
from openpyxl.styles import Protection
from openpyxl.styles import PatternFill
from glob import glob
from shutil import copyfile
from cta import update_table
#from discharge_extract import extract_specific_tags_df
from discharge_ncs import discharge_ncs
import numpy as np
from collections import defaultdict
from ActiveLearner import ActiveLearner, DISCHARGEFilter
#from featureSelection import featureSelection
from openpyxl.utils import get_column_letter

sys.path.append('H:/cloud/cloud_data/Projects/DL/Code/src')
sys.path.append('H:/cloud/cloud_data/Projects/DL/Code/src/ct')
from CTDataStruct import CTPatient
import keyboard
from sklearn.metrics import confusion_matrix
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.ensemble import RandomForestClassifier
from numpy.random import shuffle
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule


def isfloat(value):
  try:
    float(value)
    return True
  except ValueError:
    return False

def computeCTA(settings):
    print('computeCTA')
    #folderpath_master = 'H:/cloud/cloud_data/Projects/CACSFilter/data/discharge_master/discharge_master_01042020'
    #date = folderpath_master.split('_')[-1]
    #folderpath_components = os.path.join(folderpath_master, 'discharge_components_' + date)
    #folderpath_sources = os.path.join(folderpath_master, 'discharge_sources_' + date)
    #filepath_master = os.path.join(folderpath_master, 'discharge_master_' + date + '.xlsx')
    #filepath_data = os.path.join(folderpath_components, 'discharge_data_' + date + '.xlsx')   
    
    filepath_dicom = settings['filepath_dicom']
    filepath_master = settings['filepath_data']
    filepath_ITT = settings['filepath_ITT']
    filepath_phase_exclude_stenosis = settings['filepath_phase_exclude_stenosis']
    filepath_stenosis_bigger_20_phases = settings['filepath_stenosis_bigger_20_phases']
    filepath_prct = settings['filepath_prct']
    filepath_ecrf = settings['filepath_ecrf']
    #filepath_master = 'H:/cloud/cloud_data/Projects/CACSFilter/data/discharge_master/discharge_master_01042020/discharge_master_01042020.xlsx'
    #filepath_master = 'H:/cloud/cloud_data/Projects/CACSFilter/data/discharge_master/discharge_master_01042020/discharge_master_01042020.xlsx'
    
    df_discharge = pd.read_excel(filepath_dicom)
    df_master = pd.read_excel(filepath_master)
    df_ITT = pd.read_excel(filepath_ITT)
    df_phase_exclude_stenosis = pd.read_excel(filepath_phase_exclude_stenosis)
    df_stenosis_bigger_20_phase = pd.read_excel(filepath_stenosis_bigger_20_phases)
    df_ITT = pd.read_excel(filepath_ITT)
    df_prct = pd.read_excel(filepath_prct)
    df_ecrf = pd.read_excel(filepath_ecrf)
    
    # 34
    df_ecrf_tmp = df_ecrf[df_ecrf['1. Date of CT scan'].notna()]
    df_ITT_tmp = df_ITT[~(df_ITT['ITT']==2)]
    ct_ecrf = pd.merge(df_ecrf_tmp, df_ITT_tmp, left_on='Patient identifier', right_on='ID')[['Patient identifier', '1. Date of CT scan']].drop_duplicates('Patient identifier')
    ct_ecrf = ct_ecrf.rename(columns={'1. Date of CT scan': 'EcrfDate'})
    
    # 37
    df_phase_exclude_stenosis_tmp = df_phase_exclude_stenosis[df_phase_exclude_stenosis['phase_i0011'].notna()]
    phase_no_stenosis = df_phase_exclude_stenosis_tmp[['mnpaid', 'phase_i0011']]
    phase_no_stenosis = phase_no_stenosis.rename(columns={'mnpaid': 'PatientID', 'phase_i0011': 'phase'})
    phase_no_stenosis['arteries'] = 'LAD,RCA,LCX,LMA'
    
    #40
    # create view phase_with_stenosis1 as select mnpaid as PatientID, `sten_i0231 (Phase #1)` as phase, concat_ws(',',if(LAD=1,'LAD', null),if(RCA=1,'RCA',null),if(LMA=1,'LMA',null),if(LCX=1,'LCX',null)) as arteries from stenosis_bigger_20_phases where `sten_i0231 (Phase #1)` is not null;
    df_stenosis_bigger_20_phase_tmp = df_stenosis_bigger_20_phase[df_stenosis_bigger_20_phase['sten_i0231 (Phase #1)'].notna()]
    df_stenosis_bigger_20_phase_tmp = df_stenosis_bigger_20_phase_tmp.reset_index(drop=True)
    phase_with_stenosis1 = df_stenosis_bigger_20_phase_tmp[['mnpaid', 'sten_i0231 (Phase #1)']]
    arteries_tmp=pd.DataFrame('', index=np.arange(len(phase_with_stenosis1)), columns=['arteries']) 
    for index, row in arteries_tmp.iterrows():
        s=''
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LAD']==1:
            s = s + ',LAD'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'RCA']==1:
            s = s + ',RCA'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LMA']==1:
            s = s + ',LMA'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LCX']==1:
            s = s + ',LCX'
        if len(s)==0:
            arteries_tmp.loc[index,'arteries'] = np.nan
        else:
            arteries_tmp.loc[index,'arteries'] = s[1:]
    phase_with_stenosis1['arteries'] = arteries_tmp
    phase_with_stenosis1 = phase_with_stenosis1.rename(columns={'mnpaid': 'PatientID', 'sten_i0231 (Phase #1)': 'phase'})
    
    # 41
    df_stenosis_bigger_20_phase_tmp = df_stenosis_bigger_20_phase[df_stenosis_bigger_20_phase['sten_i0241'].notna()]
    df_stenosis_bigger_20_phase_tmp = df_stenosis_bigger_20_phase_tmp.reset_index(drop=True)
    phase_with_stenosis2 = df_stenosis_bigger_20_phase_tmp[['mnpaid', 'sten_i0241']]
    arteries_tmp=pd.DataFrame('', index=np.arange(len(phase_with_stenosis2)), columns=['arteries']) 
    for index, row in arteries_tmp.iterrows():
        s=''
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LAD']==1:
            s = s + ',LAD'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'RCA']==1:
            s = s + ',RCA'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LMA']==1:
            s = s + ',LMA'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LCX']==1:
            s = s + ',LCX'
        if len(s)==0:
            arteries_tmp.loc[index,'arteries'] = np.nan
        else:
            arteries_tmp.loc[index,'arteries'] = s[1:]
    phase_with_stenosis2['arteries'] = arteries_tmp
    phase_with_stenosis2 = phase_with_stenosis2.rename(columns={'mnpaid': 'PatientID', 'sten_i0241': 'phase'})
    
    # 42
    df_stenosis_bigger_20_phase_tmp = df_stenosis_bigger_20_phase[df_stenosis_bigger_20_phase['sten_i0251'].notna()]
    df_stenosis_bigger_20_phase_tmp = df_stenosis_bigger_20_phase_tmp.reset_index(drop=True)
    phase_with_stenosis3 = df_stenosis_bigger_20_phase_tmp[['mnpaid', 'sten_i0251']]
    arteries_tmp=pd.DataFrame('', index=np.arange(len(phase_with_stenosis3)), columns=['arteries']) 
    for index, row in arteries_tmp.iterrows():
        s=''
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LAD']==1:
            s = s + ',LAD'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'RCA']==1:
            s = s + ',RCA'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LMA']==1:
            s = s + ',LMA'
        if df_stenosis_bigger_20_phase_tmp.loc[index,'LCX']==1:
            s = s + ',LCX'
        if len(s)==0:
            arteries_tmp.loc[index,'arteries'] = np.nan
        else:
            arteries_tmp.loc[index,'arteries'] = s[1:]
    phase_with_stenosis3['arteries'] = arteries_tmp
    phase_with_stenosis3 = phase_with_stenosis3.rename(columns={'mnpaid': 'PatientID', 'sten_i0251': 'phase'})
    print('computeCTA01')
    # 43
    # create view phase_information as select * from  phase_no_stenosis union select * from phase_with_stenosis1 union select * from phase_with_stenosis2 union select * from phase_with_stenosis3;
    phase_information = pd.concat([phase_no_stenosis, phase_with_stenosis1, phase_with_stenosis2, phase_with_stenosis3], axis=0).drop_duplicates()
    
    
    # 47
    # create view rca_double as select PatientID, group_concat(distinct phase), count(distinct phase) from phase_information where instr(arteries,'RCA') group by PatientID having count(distinct phase)>1;
    phase_information_tmp = phase_information.replace(to_replace=[np.nan], value='', inplace=False)
    phase_information_tmp = phase_information_tmp[phase_information_tmp['arteries'].str.contains('RCA')]
    rca_double=pd.DataFrame(columns=['PatientID', 'group_concat(distinct phase)', 'count(distinct phase)'])
    patients = phase_information_tmp['PatientID'].unique()
    # Error for 29-TUR-0005 (LAD,RCA,LCX), (RCA) ???
    for pat in patients:
        df_pat = phase_information_tmp[phase_information_tmp['PatientID']==pat]
        s=''
        for index, row in df_pat.iterrows():
            s = s + ',' + str(int(row['phase']))
        if len(s)>0:
            s = s[1:]
        count=df_pat['arteries'].value_counts().max()
        
        if count>1:
            rca_double = rca_double.append({'PatientID':pat, 'group_concat(distinct phase)':s, 'count(distinct phase)':count}, ignore_index=True)
    
    # 48
    phase_information_tmp = phase_information.replace(to_replace=[np.nan], value='', inplace=False)
    phase_information_tmp = phase_information_tmp[phase_information_tmp['arteries'].str.contains('LMA')]
    lma_double=pd.DataFrame(columns=['PatientID', 'group_concat(distinct phase)', 'count(distinct phase)'])
    patients = phase_information_tmp['PatientID'].unique()
    # Error for 29-TUR-0005 (LAD,RCA,LCX), (RCA) ???
    for pat in patients:
        df_pat = phase_information_tmp[phase_information_tmp['PatientID']==pat]
        s=''
        for index, row in df_pat.iterrows():
            s = s + ',' + str(int(row['phase']))
        if len(s)>0:
            s = s[1:]
        count=df_pat['arteries'].value_counts().max()
        
        if count>1:
            lma_double = lma_double.append({'PatientID':pat, 'group_concat(distinct phase)':s, 'count(distinct phase)':count}, ignore_index=True)
    
    # 49
    phase_information_tmp = phase_information.replace(to_replace=[np.nan], value='', inplace=False)
    phase_information_tmp = phase_information_tmp[phase_information_tmp['arteries'].str.contains('LAD')]
    lad_double=pd.DataFrame(columns=['PatientID', 'group_concat(distinct phase)', 'count(distinct phase)'])
    patients = phase_information_tmp['PatientID'].unique()
    # Error for 29-TUR-0005 (LAD,RCA,LCX), (RCA) ???
    for pat in patients:
        df_pat = phase_information_tmp[phase_information_tmp['PatientID']==pat]
        s=''
        for index, row in df_pat.iterrows():
            s = s + ',' + str(int(row['phase']))
        if len(s)>0:
            s = s[1:]
        count=df_pat['arteries'].value_counts().max()
        
        if count>1:
            lad_double = lad_double.append({'PatientID':pat, 'group_concat(distinct phase)':s, 'count(distinct phase)':count}, ignore_index=True)
    
    # 50
    phase_information_tmp = phase_information.replace(to_replace=[np.nan], value='', inplace=False)
    phase_information_tmp = phase_information_tmp[phase_information_tmp['arteries'].str.contains('LCX')]
    lcx_double=pd.DataFrame(columns=['PatientID', 'group_concat(distinct phase)', 'count(distinct phase)'])
    patients = phase_information_tmp['PatientID'].unique()
    # Error for 29-TUR-0005 (LAD,RCA,LCX), (RCA) ???
    for pat in patients:
        df_pat = phase_information_tmp[phase_information_tmp['PatientID']==pat]
        s=''
        for index, row in df_pat.iterrows():
            s = s + ',' + str(int(row['phase']))
        if len(s)>0:
            s = s[1:]
        count=df_pat['arteries'].value_counts().max()
        
        if count>1:
            lcx_double = lcx_double.append({'PatientID':pat, 'group_concat(distinct phase)':s, 'count(distinct phase)':count}, ignore_index=True)
    
    print('computeCTA02')
    # 51
    # create view phase_double as select *, 'LAD' as vessel from lad_double union select *, 'RCA' from rca_double union select *, 'LMA' from lma_double union select *, 'LCX' from lcx_double order by PatientID;
    phase_double = pd.concat([rca_double, lma_double, lad_double, lcx_double], axis=0).copy()
    phase_double= phase_double.reset_index(drop=True)
    vessel = pd.Series(['RCA' for i in range(len(rca_double))] + ['LMA' for i in range(len(lma_double))] + ['LAD' for i in range(len(lad_double))] +['LCX' for i in range(len(lcx_double))])
    phase_double['vessel'] = vessel
    phase_double = phase_double.sort_values('PatientID')
    
    # 54
    # create view phase_oka as select * from phase_information where PatientID not in (select distinct PatientID from phase_double) order by PatientID;
    phase_information = phase_information.reset_index(drop=True)
    phase_double = phase_double.reset_index(drop=True)
    phase_oka = pd.DataFrame(columns=phase_double.columns)
    patients = list(phase_double['PatientID'])
    for index, row in phase_information.iterrows():
        if not row['PatientID'] in patients:
            phase_oka = phase_oka.append(row)
    phase_oka = phase_oka[['PatientID', 'phase', 'arteries']]
    # 56
    phase_ok = phase_oka.copy()
    
    # 59
    df_prct_tmp = df_prct[df_prct['other_best_phase'].notna()]
    prct_phase_other = df_prct_tmp[['PatientId', 'other_best_phase']]
    prct_phase_other['arteries'] = 'RCA, LAD, LCX'
    prct_phase_other = prct_phase_other.rename(columns={'other_best_phase': 'phase'})
    
    # 60
    df_prct_tmp = df_prct[df_prct['rca_best_phase'].notna()]
    prct_phase_rca = df_prct_tmp[['PatientId', 'rca_best_phase']]
    prct_phase_rca['arteries'] = 'RCA'
    prct_phase_rca = prct_phase_rca.rename(columns={'rca_best_phase': 'phase'})
    
    # 61
    df_prct_tmp = df_prct[df_prct['lad_best_phase'].notna()]
    prct_phase_lad = df_prct_tmp[['PatientId', 'lad_best_phase']]
    prct_phase_lad['arteries'] = 'LAD'
    prct_phase_lad = prct_phase_lad.rename(columns={'lad_best_phase': 'phase'})
    
    
    # 62
    df_prct_tmp = df_prct[df_prct['lcx_best_phase'].notna()]
    prct_phase_lcx = df_prct_tmp[['PatientId', 'lcx_best_phase']]
    prct_phase_lcx['arteries'] = 'LCX'
    prct_phase_lcx = prct_phase_lcx.rename(columns={'lcx_best_phase': 'phase'})
    
    #63
    #prct_phases = pd.concat([prct_phase_other, prct_phase_rca, prct_phase_lad, prct_phase_lcx], axis=0).drop_duplicates() 
    # Replaced to filter phase which ar strings (comments)
    prct_phases_tmp = pd.concat([prct_phase_other, prct_phase_rca, prct_phase_lad, prct_phase_lcx], axis=0).drop_duplicates() 
    prct_phases = pd.DataFrame(columns=prct_phases_tmp.columns)
    for index, row in prct_phases_tmp.iterrows():
        if isfloat(row['phase']):
            prct_phases = prct_phases.append(row)
    
    #66
    #create view rca_double as select PatientID, group_concat(distinct phase), count(distinct phase) from phase_information where instr(arteries,'RCA') group by PatientID having count(distinct phase)>1;
    #create view prct_rca_double as select PatientId, group_concat(distinct phase), count(distinct phase) from prct_phases where instr(arteries, 'RCA') group by PatientId having count(distinct phase)>1; 
    prct_phases_tmp = prct_phases.replace(to_replace=[np.nan], value='', inplace=False)
    prct_phases_tmp = prct_phases_tmp[prct_phases_tmp['arteries'].str.contains('RCA')]
    prct_rca_double=pd.DataFrame(columns=['PatientID', 'group_concat(distinct phase)', 'count(distinct phase)'])
    patients = prct_phases_tmp['PatientId'].unique()
    for pat in patients:
        if len(pat)>0:
            df_pat = prct_phases_tmp[prct_phases_tmp['PatientId']==pat]
            s=''
            for index, row in df_pat.iterrows():
                s = s + ',' + row['phase']
            if len(s)>0:
                s = s[1:]
            #count=df_pat['phase'].value_counts().max()
            #count=len(df_pat)
            #count=df_pat['arteries'].value_counts().max()
            count=len(df_pat['phase'].unique())
            
            if count>1:
                prct_rca_double = prct_rca_double.append({'PatientID':pat, 'group_concat(distinct phase)':s, 'count(distinct phase)':count}, ignore_index=True)
        else:
            prct_rca_double = prct_rca_double.append({'PatientID':pat, 'group_concat(distinct phase)':'', 'count(distinct phase)':2}, ignore_index=True)
        
    
    # 67
    prct_phases_tmp = prct_phases.replace(to_replace=[np.nan], value='', inplace=False)
    prct_phases_tmp = prct_phases_tmp[prct_phases_tmp['arteries'].str.contains('LAD')]
    prct_lad_double=pd.DataFrame(columns=['PatientID', 'group_concat(distinct phase)', 'count(distinct phase)'])
    patients = prct_phases_tmp['PatientId'].unique()
    for pat in patients:
        if len(pat)>0:
            df_pat = prct_phases_tmp[prct_phases_tmp['PatientId']==pat]
            s=''
            for index, row in df_pat.iterrows():
                s = s + ',' + row['phase']
            if len(s)>0:
                s = s[1:]
            count=len(df_pat['phase'].unique())
            
            if count>1:
                prct_lad_double = prct_lad_double.append({'PatientID':pat, 'group_concat(distinct phase)':s, 'count(distinct phase)':count}, ignore_index=True)
        else:
            prct_lad_double = prct_lad_double.append({'PatientID':pat, 'group_concat(distinct phase)':'', 'count(distinct phase)':2}, ignore_index=True)
        
        
    # 68
    prct_phases_tmp = prct_phases.replace(to_replace=[np.nan], value='', inplace=False)
    prct_phases_tmp = prct_phases_tmp[prct_phases_tmp['arteries'].str.contains('LCX')]
    prct_lcx_double=pd.DataFrame(columns=['PatientID', 'group_concat(distinct phase)', 'count(distinct phase)'])
    patients = prct_phases_tmp['PatientId'].unique()
    for pat in patients:
        if len(pat)>0:
            df_pat = prct_phases_tmp[prct_phases_tmp['PatientId']==pat]
            s=''
            for index, row in df_pat.iterrows():
                s = s + ',' + row['phase']
            if len(s)>0:
                s = s[1:]
            count=len(df_pat['phase'].unique())
            
            if count>1:
                prct_lcx_double = prct_lcx_double.append({'PatientID':pat, 'group_concat(distinct phase)':s, 'count(distinct phase)':count}, ignore_index=True)
        else:
            prct_lcx_double = prct_lcx_double.append({'PatientID':pat, 'group_concat(distinct phase)':'', 'count(distinct phase)':2}, ignore_index=True)
        
    # 69
    prct_phase_double = pd.concat([prct_rca_double, prct_lad_double, prct_lcx_double], axis=0).copy()
    prct_phase_double = prct_phase_double.reset_index(drop=True)
    prct_phase_double = prct_phase_double.sort_values('PatientID')
    
    print('computeCTA03')
    # 72
    #create view prct_oka as select prct_phases.* from prct_phases left join  prct_phase_double on prct_phases.PatientId=prct_phase_double.PatientId where prct_phase_double.PatientId is null;
    prct_phases = prct_phases.reset_index(drop=True)
    prct_phase_double = prct_phase_double.reset_index(drop=True)
    prct_oka = pd.DataFrame(columns=prct_phases.columns)
    patients = list(prct_phase_double['PatientID'])
    for index, row in prct_phases.iterrows():
        if not row['PatientId'] in patients:
            prct_oka = prct_oka.append(row)
    
    
    # 73
    # create view prct_problematic as select PatientId, group_concat(distinct phase), group_concat(arteries), count(distinct phase) from prct_oka group by PatientId having count(distinct phase)=2 or count(distinct phase)>3;
    #for index, row in prct_oka.iterrows():
        
    # 74
    # create view prct_ok as select prct_oka.* from prct_oka;   
    prct_ok = prct_oka.copy()
    
    # 77
    # create table useful_ecrf_phases as select phase_ok.*, ct_ecrf.EcrfDate from phase_ok left join ct_ecrf on PatientID=`Patient identifier` where `Patient identifier` is not null;    
    useful_ecrf_phases = pd.merge(phase_ok, ct_ecrf, left_on='PatientID', right_on='Patient identifier')
    useful_ecrf_phases = useful_ecrf_phases[['PatientID', 'phase', 'arteries', 'EcrfDate']]
    
    # 78
    # create table ct_ecrf_no_admissible_phase as select ct_ecrf.* from ct_ecrf left join useful_ecrf_phases on PatientID=`Patient identifier` where PatientID is null;
    ct_ecrf_no_admissible_phase = pd.DataFrame(columns=ct_ecrf.columns)
    patients = list(useful_ecrf_phases['PatientID'])
    for index, row in ct_ecrf.iterrows():
        if not row['Patient identifier'] in patients:
            ct_ecrf_no_admissible_phase = ct_ecrf_no_admissible_phase.append(row)
    # 79
    # create table useful_prct_phases as select prct_ok.*, ct_ecrf.EcrfDate from prct_ok left join ct_ecrf on PatientId=`Patient identifier`;
    useful_prct_phases = pd.merge(prct_ok, ct_ecrf, left_on='PatientId', right_on='Patient identifier', how='left')
    useful_prct_phases=useful_prct_phases[['PatientId', 'phase', 'arteries', 'EcrfDate']]
    
    # 80
    # create table manual as select ct_ecrf_no_admissible_phase.* from ct_ecrf_no_admissible_phase left join useful_prct_phases on `Patient identifier`=PatientId where PatientId is null;
    manual = pd.DataFrame(columns=ct_ecrf_no_admissible_phase.columns)
    patients = list(useful_prct_phases['PatientId'])
    for index, row in ct_ecrf_no_admissible_phase.iterrows():
        if not row['Patient identifier'] in patients:
            manual = manual.append(row)
    
    ###########################################################################################
    
    def isNaN(num):
        return num != num
    
    admissible = df_master[['PatientID', 'SeriesInstanceUID', 'SeriesDescription', 'SliceThickness', 'Modality',
                            'Rows', 'ImageComments', 'NominalPercentageOfCardiacPhase', 'CardiacRRIntervalSpecified',
                            'StudyDate', 'ConvolutionKernel']]
    
    kernel = 'B35s|Qr36d|FC12|FC51|FC17|B60f|B70f|B30f|B31f|B08s|B19f|B20f|B20s|B30s|B31s|B40f|B41s|B50f|B50s|B65f|B70f|B70s|B80s|Bf32dB80s|Bf32d|Bl57d|Br32d|Bv36d|Bv40f|FC08|FC08-H|FC15|FC18|FC35|FC52|FL03|FL04|FL05|H20f|H31s|IMR1|IMR2|IMR2|Qr36d|T20f|T20s|Tr20f|UB|XCA|YA'
    
    print('computeCTA04')
    bool_Rows = (df_master['Rows'].isna()) | (df_master['Rows']<600)
    bool_Modality = (df_master['Modality']=='CT') | (df_master['Modality']=='OT')
    bool_SliceThickness = ((df_master['SliceThickness']>0) & (df_master['SliceThickness']<0.8)) | (df_master['SliceThickness'].isna())
    bool_kernel = bool_SliceThickness.copy()
    for index, row in df_master.iterrows():
        if isNaN(row['ConvolutionKernel']):
            bool_kernel[index] = True
        else:
            bool_kernel[index] = not row['ConvolutionKernel'] in kernel
            
    bool_series = bool_SliceThickness.copy()
    for index, row in df_master.iterrows():
        if isNaN(row['SeriesDescription']):
            bool_series[index] = True
        else:
            bool_series[index] = not row['SeriesDescription'] in kernel
    
    bool_pixel = bool_SliceThickness.copy()
    for index, row in df_master.iterrows():
        if isNaN(row['PixelSpacing']):
            bool_pixel[index] = True
        else:
            PixelSpacing = float(row['PixelSpacing'].split(',')[0][1:])
            Rows = float(row['Rows'])
            bool_pixel[index] = PixelSpacing*Rows < 260
    
    idx = bool_Rows & bool_Modality & bool_SliceThickness & bool_kernel & bool_series & bool_pixel
    admissible = admissible[idx]
    admissible = admissible.rename(columns={'PatientID': 'jabroni'})
    
    # 33
    # create view really_useful_ecrf as select PatientID, group_concat(distinct phase) as phase, group_concat(distinct arteries) as arteries, group_concat(distinct EcrfDate) as ecrf_date from useful_ecrf_phases group by PatientID, phase;
    useful_ecrf_phases_tmp = useful_ecrf_phases.replace(to_replace=[np.nan], value='', inplace=False)
    really_useful_ecrf = pd.DataFrame(columns=useful_ecrf_phases_tmp.columns)
    patients = list(useful_ecrf_phases_tmp['PatientID'].unique())
    for patient in patients:
        df_pat = useful_ecrf_phases_tmp[useful_ecrf_phases_tmp['PatientID']==patient]
        phases = list(df_pat['phase'].unique())
        #phases = [int(x) for x in phasees]
    
        for phase in phases:
            df_phase = df_pat[df_pat['phase'] == phase]
            arteries=''
            for index, row in df_phase.iterrows():
                if arteries=='':
                    arteries = arteries + row['arteries']
                else:
                    if not row['arteries'] in arteries:
                        arteries = arteries + ',' + row['arteries']
                    
            really_useful_ecrf = really_useful_ecrf.append({'PatientID': patient, 'phase': phase, 'arteries': arteries,'EcrfDate': row['EcrfDate']}, ignore_index=True)
    really_useful_ecrf = really_useful_ecrf.rename(columns={'EcrfDate': 'ecrf_date'})
    
    # 34
    useful_prct_phases_tmp = useful_prct_phases.replace(to_replace=[np.nan], value='', inplace=False)
    really_useful_prct = pd.DataFrame(columns=useful_prct_phases_tmp.columns)
    patients = list(useful_prct_phases_tmp['PatientId'].unique())
    for patient in patients:
        df_pat = useful_prct_phases_tmp[useful_prct_phases_tmp['PatientId']==patient]
        phases = list(df_pat['phase'].unique())
        #phases = [int(x) for x in phasees]
    
        for phase in phases:
            df_phase = df_pat[df_pat['phase'] == phase]
            arteries=''
            for index, row in df_phase.iterrows():
                if arteries=='':
                    arteries = arteries + row['arteries']
                else:
                    if not row['arteries'] in arteries:
                        arteries = arteries + ',' + row['arteries']
                    
            really_useful_prct = really_useful_prct.append({'PatientId': patient, 'phase': phase, 'arteries': arteries,'EcrfDate': row['EcrfDate']}, ignore_index=True)
    really_useful_prct = really_useful_prct.rename(columns={'EcrfDate': 'ecrf_date'})
    
    # 37
    # create view ecrf_dicom_phases as select PatientID, SeriesDescription,SeriesInstanceUID,ecrf_date, phase,arteries, Modality, ImageComments, admissible.Rows from really_useful_ecrf left join admissible on PatientID=jabroni where SeriesInstanceUID is not null and (instr(SeriesDescription,phase) or SeriesDescription is null or convert(phase, double) between NominalPercentageOfCardiacPhase -1 and NominalPercentageOfCardiacPhase +1);
    ecrf_dicom_phases_tmp = pd.merge(really_useful_ecrf, admissible, left_on='PatientID', right_on='jabroni')
    ecrf_dicom_phases_tmp = ecrf_dicom_phases_tmp.replace(to_replace=[np.nan], value='', inplace=False)
    ecrf_dicom_phases = pd.DataFrame(columns=ecrf_dicom_phases_tmp.columns)
    for index, row in ecrf_dicom_phases_tmp.iterrows():
        c0 = str(int(row['phase'])) in row['SeriesDescription']
        c1 = row['SeriesDescription'] == ''
        if not row['NominalPercentageOfCardiacPhase']=='':
            nom_min = float(row['NominalPercentageOfCardiacPhase'])-1
            nom_max = float(row['NominalPercentageOfCardiacPhase'])+1
            c2 = nom_min < float(row['phase']) < nom_max
        else:
            c2 = False
        if c0 or c1 or c2:
            ecrf_dicom_phases = ecrf_dicom_phases.append(row.copy())
    ecrf_dicom_phases = ecrf_dicom_phases[['PatientID', 'SeriesInstanceUID', 'SeriesDescription', 'ecrf_date', 'phase', 'arteries', 'Modality', 'ImageComments', 'Rows']]
    ecrf_dicom_phases = ecrf_dicom_phases.sort_values('PatientID') 
    
    print('computeCTA05')
    # 39
    # create view prct_dicom_phases as select PatientId,  SeriesDescription,SeriesInstanceUID,ecrf_date, phase,arteries, Modality, ImageComments, admissible.Rows from really_useful_prct left join admissible on PatientId=jabroni where SeriesInstanceUID is not null and (SeriesDescription REGEXP concat(phase-2,"|",phase-1,"|",phase,"|",phase+1,"|",phase+2) or SeriesDescription REGEXP concat(round(phase*CardiacRRIntervalSpecified*0.01)-2,"|", round(phase*CardiacRRIntervalSpecified*0.01)-1,"|", round(phase*CardiacRRIntervalSpecified*0.01),"|", round(phase*CardiacRRIntervalSpecified*0.01)+1,"|", round(phase*CardiacRRIntervalSpecified*0.01)+2) or SeriesDescription is null or convert(phase, double) between NominalPercentageOfCardiacPhase -1 and NominalPercentageOfCardiacPhase +1);
    prct_dicom_phases_tmp = pd.merge(really_useful_prct, admissible, left_on='PatientId', right_on='jabroni')
    prct_dicom_phases_tmp = prct_dicom_phases_tmp.replace(to_replace=[np.nan], value='', inplace=False)
    prct_dicom_phases = pd.DataFrame(columns=prct_dicom_phases_tmp.columns)
    for index, row in prct_dicom_phases_tmp.iterrows():
        phase = int(round(float(row['phase'])))
        # if row['PatientId']=='01-BER-0045':
        #     pass#sys.exit()
        # if isfloat(row['phase']):
        #     c0 = str(phase-2) in row['SeriesDescription'] or str(phase-1) in row['SeriesDescription'] or str(phase) in row['SeriesDescription'] or str(phase+1) in row['SeriesDescription'] or str(phase+2) in row['SeriesDescription']
        # else:
        #     r0 = -2
        #     r1 = -1
        #     #r2 = 0
        #     r3 = 1
        #     r4 = 2
        #     c0 = str(r0) in row['SeriesDescription'] or str(r1) in row['SeriesDescription'] or str(r3) in row['SeriesDescription'] or str(r4) in row['SeriesDescription']
        
        c0 = str(phase-2) in row['SeriesDescription'] or str(phase-1) in row['SeriesDescription'] or str(phase) in row['SeriesDescription'] or str(phase+1) in row['SeriesDescription'] or str(phase+2) in row['SeriesDescription']
        if not row['CardiacRRIntervalSpecified']=='' and isfloat(row['phase']):
            v0 = round(float(row['phase'])*float(row['CardiacRRIntervalSpecified'])*0.01)-2
            v1 = round(float(row['phase'])*float(row['CardiacRRIntervalSpecified'])*0.01)-1
            v2 = round(float(row['phase'])*float(row['CardiacRRIntervalSpecified'])*0.01)
            v3 = round(float(row['phase'])*float(row['CardiacRRIntervalSpecified'])*0.01)+1
            v4 = round(float(row['phase'])*float(row['CardiacRRIntervalSpecified'])*0.01)+2
            c1 = str(v0) in row['SeriesDescription'] or str(v1) in row['SeriesDescription'] or str(v2) in row['SeriesDescription'] or str(v3) in row['SeriesDescription'] or str(v4) in row['SeriesDescription']
        else:
            c1 = False
        if not row['NominalPercentageOfCardiacPhase']=='':
            nom_min = float(row['NominalPercentageOfCardiacPhase'])-1
            nom_max = float(row['NominalPercentageOfCardiacPhase'])+1
            c2 = nom_min < float(row['phase']) < nom_max
        else:
            c2 = False
        if row['SeriesDescription']=='':
            c3 = True
        else:
            c3 = False
        if c0 or c1 or c2 or c3:
            prct_dicom_phases = prct_dicom_phases.append(row.copy())
    prct_dicom_phases = prct_dicom_phases[['PatientId', 'SeriesInstanceUID', 'SeriesDescription', 'ecrf_date', 'phase', 'arteries', 'Modality', 'ImageComments', 'Rows']]
    prct_dicom_phases = prct_dicom_phases.sort_values('PatientId') 
    
    
    # 47
    # create view no_prct_dicom_phases as select useful_prct_phases.* from useful_prct_phases left join prct_dicom_phases on useful_prct_phases.PatientId=prct_dicom_phases.PatientId where prct_dicom_phases.PatientId is null;
    prct_not_ecrf_dicom_phases = pd.DataFrame(columns=prct_dicom_phases.columns) 
    patients = list(ecrf_dicom_phases['PatientID'].unique())
    for index, row in prct_dicom_phases.iterrows():
        if not row['PatientId'] in patients:
            prct_not_ecrf_dicom_phases = prct_not_ecrf_dicom_phases.append(row.copy())
    
    # 53
    # create view phases as select SeriesInstanceUID, ecrf_date, phase,arteries, "ecrf" as source from ecrf_dicom_phases union select SeriesInstanceUID, ecrf_date, phase, arteries, "prct" from prct_not_ecrf_dicom_phases;
    
    prct_not_ecrf_dicom_phases_tmp = prct_not_ecrf_dicom_phases.rename(columns={'PatientId': 'PatientID'})
    phases_tmp = pd.concat([ecrf_dicom_phases, prct_not_ecrf_dicom_phases_tmp], axis=0)
    phases_tmp = phases_tmp.reset_index(drop=True)
    source = pd.concat([pd.DataFrame('ecrf', index=np.arange(len(ecrf_dicom_phases)), columns=['source']), 
                        pd.DataFrame('prct', index=np.arange(len(prct_not_ecrf_dicom_phases_tmp)), columns=['source'])], axis=0)
    source = source.reset_index(drop=True)
    phases = pd.concat([phases_tmp, source],axis=1)
    phases = phases[['SeriesInstanceUID', 'ecrf_date', 'phase', 'arteries', 'source']]
    
    # 54
    # create view admissible_phases as select admissible.*,ecrf_date, phase, arteries, source from admissible left join phases on phases.SeriesInstanceUID=admissible.SeriesInstanceUID;
    admissible_phases = pd.merge(admissible, phases, left_on='SeriesInstanceUID', right_on='SeriesInstanceUID', how='left')
    
    
    admissible = df_master[['PatientID', 'SeriesInstanceUID', 'SeriesDescription', 'SliceThickness', 'Modality',
                            'Rows', 'ImageComments', 'NominalPercentageOfCardiacPhase', 'CardiacRRIntervalSpecified',
                            'StudyDate', 'ConvolutionKernel']]
    # 58
    master_haha = df_master.copy()
    bool_Rows = (master_haha['Rows'].isna()) | (master_haha['Rows']<600)
    bool_Modality = (master_haha['Modality']=='CT') | (master_haha['Modality']=='OT')
    bool_SliceThickness = ((master_haha['SliceThickness']>0) & (master_haha['SliceThickness']<0.8)) | (master_haha['SliceThickness'].isna())
    
    print('computeCTA06')
    kernel = 'B35s|Qr36d|FC12|FC51|FC17|B60f|B70f|B30f|B31f|B08s|B19f|B20f|B20s|B30s|B31s|B40f|B41s|B50f|B50s|B65f|B70f|B70s|B80s|Bf32dB80s|Bf32d|Bl57d|Br32d|Bv36d|Bv40f|FC08|FC08-H|FC15|FC18|FC35|FC52|FL03|FL04|FL05|H20f|H31s|IMR1|IMR2|IMR2|Qr36d|T20f|T20s|Tr20f|UB|XCA|YA'
    bool_kernel = bool_SliceThickness.copy()
    for index, row in master_haha.iterrows():
        print(index)
        if isNaN(row['ConvolutionKernel']):
            bool_kernel[index] = True
        else:
            bool_kernel[index] = not row['ConvolutionKernel'] in kernel
            
    print('computeCTA07')
            
    bool_series = bool_SliceThickness.copy()
    for index, row in master_haha.iterrows():
        print(index)
        if isNaN(row['SeriesDescription']):
            bool_series[index] = True
        else:
            bool_series[index] = not row['SeriesDescription'] in kernel
    
    bool_pixel = bool_SliceThickness.copy()
    for index, row in master_haha.iterrows():
        print(index)
        if isNaN(row['PixelSpacing']):
            bool_pixel[index] = True
        else:
            PixelSpacing = float(row['PixelSpacing'].split(',')[0][1:])
            Rows = float(row['Rows'])
            bool_pixel[index] = PixelSpacing*Rows < 260

    
    bool_count = bool_SliceThickness.copy()
    for index, row in master_haha.iterrows():
        print(index)
        bool_count[index] = row['Count']>100
            
    CTA = bool_Rows & bool_Modality & bool_SliceThickness & bool_kernel & bool_series & bool_pixel & bool_count
    
    master_plus_ecrf = master_haha.copy()
    for index, row in master_plus_ecrf.iterrows():
        print(index)
        df = admissible_phases[admissible_phases['SeriesInstanceUID']==row['SeriesInstanceUID']]
        if df.shape[0]==1:
            master_plus_ecrf.loc[index, 'phase'] = list(df['phase'])[0]
            master_plus_ecrf.loc[index, 'arteries'] = list(df['arteries'])[0]
            master_plus_ecrf.loc[index, 'source'] = list(df['source'])[0]
    master_plus_ecrf['CTA'] = CTA
    master_plus_ecrf = master_plus_ecrf.sort_values(['PatientID', 'SeriesInstanceUID']) 
    master_plus_ecrf = master_plus_ecrf.rename(columns={'phase': 'CTA_phase', 'arteries': 'CTA_arteries', 'source': 'CTA_source'})
    
    print('computeCTA out')
    return master_plus_ecrf
    
    
    
    






