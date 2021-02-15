# -*- coding: utf-8 -*-
import sys, os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
from collections import defaultdict
from scanner_map import searchKey, CertifiedManufacturerModelNameCTDict, CertifiedManufacturerCTDict, TrueManufacturerModelNameCTDict, TrueManufacturerCTDict
from scanner_map import ScannerType, CertifiedManufacturerModelNameICADict, CertifiedManufacturerICADict, TrueManufacturerModelNameICADict, TrueManufacturerICADict
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

def highlight_columns(sheet, columns=[], color='A5A5A5', offset=2):
    for col in columns:
        cell = sheet.cell(1, col+offset)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type = 'solid')
    return sheet  

def merge_defaultdicts(d,d1):
    for k,v in d1.items():
        if (k in d):
            d[k].update(d1[k])
        else:
            d[k] = d1[k]
    return d

def covertDate(date_str):
    month_lookup = defaultdict(lambda: None, {'JAN':1, 'FEB':2, 'MAR':3, 'APR':4, 'MAY':5, 'JUN':6, 'JUL':7, 'AUG':8,'SEP':9, 'OCT':10,'NOV':11, 'DEC':12})
    day = str(date_str[0:2])
    month = str(month_lookup[date_str[2:5]])
    year = date_str[5:9]
    s = year + month + day
    return datetime.strptime(s, '%Y%m%d')

def checkModalities(modList0, modList1):
    for m0 in modList0:
        for m1 in modList1:
            if m0==m1:
                return True
    return False


def splitScannerList(filepath_scanner):
    #filepath_scanner = 'H:/cloud/cloud_data/Projects/CACSFilter/data/scanner/scanner_correspondence_V05_manual.xlsx'

    df_scanner = pd.read_excel(filepath_scanner, 'linear', index_col=0)
    df_missing_CT = pd.DataFrame(columns=df_scanner.columns)
    df_missing_XA = pd.DataFrame(columns=df_scanner.columns)

    df_missing = df_scanner[(df_scanner['ECRF_MISSING']==True) & (df_scanner['ITT']!=2)]
    for index, row in df_missing.iterrows():
        if 'DICOM XA' in row['ManualCheck']:
            df_missing_XA = df_missing_XA.append(row)
        if 'DICOM CT' in row['ManualCheck']:
            df_missing_CT = df_missing_CT.append(row)

    # Update CT sheet
    writer = pd.ExcelWriter(filepath_scanner, engine="openpyxl", mode="a")
    # Update CT sheet
    sheet_name = 'ECRF_MISSING_CT'
    workbook  = writer.book
    df_missing_CT.to_excel(writer, sheet_name=sheet_name, index=False)
    sheet = workbook[sheet_name]
    # Update XA sheet
    sheet_name = 'ECRF_MISSING_XA'
    workbook  = writer.book
    df_missing_XA.to_excel(writer, sheet_name=sheet_name, index=False)
    sheet = workbook[sheet_name]
    writer.save()

    
    
# Read discharge data
filepath_dicom = 'H:/cloud/cloud_data/Projects/CACSFilter/data/scanner/discharge_dicom_27082020_OT.xlsx'
filepath_ecrf_study = 'H:/cloud/cloud_data/Projects/CACSFilter/data/scanner/ecrf_study_20200827.xlsx'
filepath_scanner_old = 'H:/cloud/cloud_data/Projects/CACSFilter/data/scanner/scanner_correspondence_V04_manual.xlsx'
filepath_scanner = 'H:/cloud/cloud_data/Projects/CACSFilter/data/scanner/scanner_correspondence.xlsx'

df_dicom = pd.read_excel(filepath_dicom, 'linear', index_col=0)
#df_dicom=df_dicom[0:1000]
df_dicom.replace(to_replace=[np.nan], value='', inplace=True)
df_ecrf = pd.read_excel(filepath_ecrf_study, 'Tabelle1')
#df_ecrf=df_ecrf[0:1000]
df_scanner_old = pd.read_excel(filepath_scanner_old, 'linear', index_col=0)
df_scanner_old.replace(to_replace=[np.nan], value='', inplace=True)


columns_scanner_rename=['PatientID', 'Site', 'ITT', 'RD_MB', '1. Date of CT', 'Date of ICA scan',
                  'Date of ICA scan 2', 'Date of staged PCI 1', 'Date of staged PCI 2', 
                  'Date of staged PCI 3', 'duplicate entry', 'FFR', 'MRI_visite',
                  'Date of Echo', 'Date of PET', 'Date of SPECT:', 'Date of FU_CT-scan',
                  'Date cec_ct', 'Date pet ct', 'Date ldct', 'ldct 3m', 'ldct 6m',
                  'ldct 12m', 'Date FU ICA scan']

columns_scanner=['PatientID', 'Site', 'ITT', 'RD_MB', 
                 '1. Date of CT', '1. Date of CT StudyInstanceUID', 
                 'Date of ICA scan', 'Date of ICA scan StudyInstanceUID', 
                 'Date of ICA scan 2', 'Date of ICA scan 2 StudyInstanceUID', 
                 'Date of staged PCI 1', 'Date of staged PCI 1 StudyInstanceUID', 
                 'Date of staged PCI 2', 'Date of staged PCI 2 StudyInstanceUID', 
                 'Date of staged PCI 3', 'Date of staged PCI 3 StudyInstanceUID', 
                 'duplicate entry', 
                 'FFR', 'FFR StudyInstanceUID',
                 'MRI_visite',
                 'Date of Echo', 'Date of Echo StudyInstanceUID', 
                 'Date of PET', 'Date of PET StudyInstanceUID',
                 'Date of SPECT:', 'Date of SPECT: StudyInstanceUID', 
                 'Date of FU_CT-scan', 'Date of FU_CT-scan StudyInstanceUID',
                 'Date cec_ct', 'Date cec_ct StudyInstanceUID',
                 'Date pet ct', 'Date pet ct StudyInstanceUID', 
                 'Date ldct', 'Date ldct StudyInstanceUID', 
                 'ldct 3m', 'ldct 3m StudyInstanceUID', 
                 'ldct 6m', 'ldct 6m StudyInstanceUID',
                 'ldct 12m', 'ldct 12m StudyInstanceUID',
                 'Date FU ICA scan', 'Date FU ICA scan StudyInstanceUID']

columns_scanner_missing = [x for x in columns_scanner if x not in columns_scanner_rename]




#columns_result = ['OK', 'DICOM_MISSING', 'ECRF_MISSING', 'DICOM_ECRF_MISMATCH']
columns_result = ['DICOM_MISSING', 'ECRF_MISSING', 'ECRF_MISSING_SeriesInstanceUID']

columns_ecrf=['Patient identifier', 'Centre name (mnpctrname)', 'ITT', 'RD_MB', '1. Date of CT', 'Date of ICA scan',
                 'Date of ICA scan 2', 'Date of staged PCI 1', 'Date of staged PCI 2', 
                 'Date of staged PCI 3', 'duplicate entry ', 'FFR', 'MRI_visite',
                 'Date of Echo', 'Date of PET', 'Date of SPECT:', 'Date of FU_CT-scan:',
                 'Date cec_ct', 'Date pet ct', 'Date ldct:', 'ldct 3m', 'ldct 6m',
                 'ldct 12m', 'Date FU ICA scan:']

dates_required = ['1. Date of CT', 'Date of ICA scan', 'Date of ICA scan 2', 'Date of staged PCI 1', 'Date of staged PCI 2', 
                  'Date of staged PCI 3']

modalities_required = defaultdict(lambda: None, {'1. Date of CT': ['CT'], 'Date of ICA scan': ['XA'], 'Date of ICA scan 2': ['XA'], 
                       'Date of staged PCI 1': ['XA'], 'Date of staged PCI 2': ['XA'], 'Date of staged PCI 3': ['XA']})

dates_sidestudy = ['FFR','Date of Echo', 'Date of PET', 'Date of SPECT:', 'Date of FU_CT-scan',
                   'Date cec_ct', 'Date pet ct', 'Date ldct', 'ldct 3m', 'ldct 6m','ldct 12m', 'Date FU ICA scan']

modalities_sidestudy = defaultdict(lambda: None, {'FFR': ['XA'], 'Date of Echo': ['US'], 'Date of PET': ['CT','PT'], 'Date of SPECT:': ['CT','NM'], 'Date of FU_CT-scan': ['CT'],
                   'Date cec_ct': ['CT'], 'Date pet ct': ['PT'], 'Date ldct': ['CT'], 'ldct 3m': ['CT'], 'ldct 6m': ['CT'],'ldct 12m': ['CT'], 
                   'Date FU ICA scan': ['XA']})

dates_all = dates_required + dates_sidestudy

# f = 'H:/cloud/cloud_data/Projects/BIOQIC/08_Research/PACSServer/date.sas7bdat'
# f = 'C:/Users/bernifoellmer/Downloads/SASVisualForecasting_sampledatasets/skinproduct_vfdemo.sas7bdat'
# db = pd.read_sas(f)
         
# Create dataframe with patient per line
df_scanner = pd.DataFrame()
df_dicom_study = df_dicom.drop_duplicates(subset=['StudyInstanceUID'], ignore_index=True)

# Convert modalities into list of modalities
df_dicom_study.reset_index(drop=True,inplace=True)
for index, row in df_dicom_study.iterrows():
    print(index)
    #sys.exit()
    df = df_dicom[df_dicom['StudyInstanceUID']==row['StudyInstanceUID']]
    modList=list(set(list(df['Modality'])))
    modList_str = ','.join(modList)
    df_dicom_study.loc[index, 'Modality'] = modList_str

df_ecrf_study = df_ecrf.rename(columns = dict(zip(columns_ecrf, columns_scanner_rename)))
df_ecrf_study = df_ecrf_study[columns_scanner_rename]

# Convert date
for ecrf_date in dates_all:     
    for index, row in df_ecrf_study.iterrows():
        date_str = df_ecrf_study.loc[index, ecrf_date]
        #print('ecrf_date', ecrf_date)
        #print('index', index)
        #print('date_str', date_str)
        if (type(date_str)==str) and (not date_str=='.'):
            df_ecrf_study.loc[index, ecrf_date] = covertDate(date_str)
            
# date_str = df_ecrf_study.loc[277, 'FFR'] 

# d=covertDate(date_str)   
            
colmax=[]
for index_ecrf, row_ecrf in df_ecrf_study.iterrows():
    #sys.exit()
    df_patient = df_dicom_study[df_dicom_study['PatientID']==row_ecrf['PatientID']]
    df_patient.sort_values('StudyDate', inplace=True)
    df_patient.reset_index(inplace=True)
    print('index_ecrf:', index_ecrf)
    s = row_ecrf[columns_scanner_rename]
    for index, row in df_patient.iterrows():
        # Filter wrong  ManufacturerModelName and  Manufacturer
        TrueManufacturerList = [row['Manufacturer']]
        TrueManufacturer_str = ','.join(list(set(TrueManufacturerList)))
        TrueManufacturer = searchKey(TrueManufacturerCTDict, TrueManufacturer_str)
        TrueManufacturerModelNameList = [row['ManufacturerModelName']]
        TrueManufacturerModelName_str = ','.join(list(set(TrueManufacturerModelNameList)))
        TrueManufacturerModelName = searchKey(TrueManufacturerModelNameCTDict, TrueManufacturerModelName_str)        
        s['StudyDate' + '_' + str(index).zfill(2)] = datetime.strptime(str(row['StudyDate']), '%Y%m%d')
        s['StudyInstanceUID' + '_' + str(index).zfill(2)] = row['StudyInstanceUID']
        s['Modality' + '_' + str(index).zfill(2)] = row['Modality']
        s['TrueManufacturer' + '_' + str(index).zfill(2)] = TrueManufacturer
        s['TrueManufacturerModelName' + '_' + str(index).zfill(2)] = TrueManufacturerModelName
        s['EcrfFound' + '_' + str(index).zfill(2)] = ''
        
        if len(s.keys())>len(colmax):
            colmax = list(s.keys())
            
    df_scanner = df_scanner.append(s, ignore_index=True, sort=True)


# Add columns_scanner_missing


# Reindex columns
# df_scanner = df_scanner[colmax]
# df_scanner['DICOM_MISSING']=False
# df_scanner['ECRF_MISSING']=False
#for x in columns_scanner_missing: df_scanner[x]=''
 
colmax2=colmax.copy()
for x in columns_scanner_missing: 
    if ' StudyInstanceUID' in x:
        #sys.exit()
        k = x[0:-17]
        idx = colmax2.index(k)
        colmax2 = colmax2[0:idx+1] + [x] + colmax2[idx+1:]

for x in columns_scanner_missing: df_scanner[x]=''
df_scanner = df_scanner[colmax2]
df_scanner['DICOM_MISSING']=False
df_scanner['ECRF_MISSING']=False
df_scanner['ManualCheck']=''
df_scanner['Solved']=''
df_scanner['EI']=''
df_scanner['MB']=''
df_scanner['BF']=''
colmax2 = colmax2 + ['DICOM_MISSING', 'ECRF_MISSING', 'ManualCheck', 'Solved', 'EI', 'MB', 'BF']

# Create color dataframe
df_scanner_color = df_scanner.copy()

# Check dates from ecrf
columns_study = [c for c in df_scanner.columns if 'StudyDate' in c]
columns_study_mod = [c for c in df_scanner.columns if 'Modality' in c]
columns_study_id = [c for c in df_scanner.columns if 'StudyInstanceUID_' in c]
columns_study_found= [c for c in df_scanner.columns if 'Found in ecrf' in c]

modalities_all = merge_defaultdicts(modalities_required , modalities_sidestudy)


for index, row in df_scanner.iterrows():
    print('index', index)
    #if index==103:
    #    sys.exit()
    dates_dicom = list(row[columns_study])
    mod_dicom = list(row[columns_study_mod])
    for k in dates_all:
        if(not pd.isnull(row[k])) and (not row[k]=='.'):
            if not row[k] in dates_dicom and k in dates_required:
                df_scanner_color.loc[index, k] = 'RED'
                df_scanner.loc[index, 'DICOM_MISSING'] = True
                #print(index, 'ECRF_MISSING')
                #sys.exit()
            else:
                #if index==844:
                #    sys.exit()
                idx = [i for i,x in enumerate(dates_dicom) if x == row[k]]
                #idx_mod = [i for i in idx if modalities_all[k] in mod_dicom[i]]
                idx_mod = [i for i in idx if checkModalities(modalities_all[k], mod_dicom[i].split(','))]
                for i in idx_mod:
                    k_study = k + ' StudyInstanceUID'
                    dicom_study = columns_study_id[i]
                    if df_scanner.loc[index, k_study] == '':
                        df_scanner.loc[index, k_study] = df_scanner.loc[index, k_study] + df_scanner.loc[index, dicom_study]
                    else:
                        # if index==103:
                        #     print('add')
                        df_scanner_color.loc[index, k] = 'RED'
                        df_scanner.loc[index, k_study] = df_scanner.loc[index, k_study] + ','  + df_scanner.loc[index, dicom_study]
                        

            
df_scanner = df_scanner[colmax2]

# # Check dates from ecrf
# columns_study = newlist = [c for c in df_scanner.columns if 'StudyDate' in c]
# columns_study_mod = newlist = [c for c in df_scanner.columns if 'Modality' in c]
# for index, row in df_scanner.iterrows():
#     print('index', index)
#     #sys.exit()
#     dates_dicom = list(row[columns_study])
#     mod_dicom = list(row[columns_study_mod])
#     for k in dates_required:
#         if not pd.isnull(row[k]):
#             if not row[k] in dates_dicom:
#                 df_scanner_color.loc[index, k] = 'RED'
#                 df_scanner.loc[index, 'ECRF_MISSING'] = True
            

# Check dates from dicom
columns_study = [c for c in df_scanner.columns if 'StudyDate' in c]
columns_study_mod = newlist = [c for c in df_scanner.columns if 'Modality' in c]
for index, row in df_scanner.iterrows():
    #if row['PatientID']=='09-BEL-0002':
    #    sys.exit()
    #sys.exit()
    print('index', index)
    columns_date_ecrf = dates_required + dates_sidestudy
    dates_ecrf = list(row[columns_date_ecrf])
    
    dates_index = [(not pd.isnull(x)) and (not x=='.') for x in dates_ecrf]
    
    mod_ecrf=[]
    for i,x in enumerate(columns_date_ecrf):
        if dates_index[i]:
            mod_ecrf.append(modalities_all[x])
        else:
            mod_ecrf.append('')
            
    #mod_ecrf = [modalities_all[x] for i,x in enumerate(columns_date_ecrf) if dates_index[i]]
    #dates_ecrf = list(row[columns_date_ecrf])
    for e, k in enumerate(columns_study):
        #if k=='StudyDate_00':
        #    sys.exit()
        if not pd.isnull(row[k]):
            if not row[k] in dates_ecrf:
                if not (row[columns_study_mod[e]].split(',') == ['OT']):
                    df_scanner_color.loc[index, k] = 'GREEN'
                    df_scanner.loc[index, 'ECRF_MISSING'] = True
                    #print('ECRF_MISSING',k)
            else:
                #if index==16:
                #    sys.exit()
                
                q = row[columns_study_mod[e]].split(',')
                idx = [i for i,x in enumerate(dates_ecrf) if x == row[k]]
                #idx_mod = [i for i in idx if mod_ecrf[i] in row[columns_study_mod[e]]]
                
                #print('mode1', row[columns_study_mod[e]].split(','))
                
                idx_mod = [i for i in idx if checkModalities(mod_ecrf[i], row[columns_study_mod[e]].split(','))]
                #print('mode', row[columns_study_mod[e]].split(','))
                if len(idx_mod)>0 and not (row[columns_study_mod[e]].split(',') == ['OT']):
                    for i in idx_mod:
                        dates_ecrf[i]
                        k_date = 'EcrfFound' + k[-3:]
                        ecrf_study = columns_date_ecrf[i]
                        if df_scanner.loc[index, k_date] == '':
                            df_scanner.loc[index, k_date] = df_scanner.loc[index, k_date] + ecrf_study
                        else:
                            if not (row[columns_study_mod[e]].split(',') == ['OT']):
                                df_scanner_color.loc[index, k] = 'GREEN'
                                df_scanner.loc[index, k_date] = df_scanner.loc[index, k_date] + ',' + ecrf_study
                else:
                    if not (row[columns_study_mod[e]].split(',') == ['OT']):
                        df_scanner_color.loc[index, k] = 'GREEN'
                        df_scanner.loc[index, 'ECRF_MISSING'] = True

# Add MISSING columns
df_scanner['MISSING'] = df_scanner['DICOM_MISSING'] | df_scanner['ECRF_MISSING']
df_scanner_color['MISSING'] = df_scanner['MISSING']

# # Check dates from dicom
# columns_study = newlist = [c for c in df_scanner.columns if 'StudyDate' in c]
# for index, row in df_scanner.iterrows():
#     #sys.exit()
#     print('index', index)
#     columns_date_ecrf = dates_required + dates_sidestudy
#     dates_ecrf = list(row[columns_date_ecrf])
#     for k in columns_study:
#         if not pd.isnull(row[k]):
#             if not row[k] in dates_ecrf:
#                 df_scanner_color.loc[index, k] = 'GREEN'
#                 df_scanner.loc[index, 'DICOM_MISSING'] = True
            

# Replace ManualCheck, Solved, EI, MB, BF
for index, row in df_scanner_old.iterrows():
    #sys.exit()
    idx = df_scanner.index[df_scanner['PatientID']==row['PatientID']].tolist()[0]
    df_scanner.loc[idx, 'ManualCheck'] = row['ManualCheck']
    df_scanner.loc[idx, 'Solved'] = row['Solved']
    df_scanner.loc[idx, 'EI'] = row['EI']
    df_scanner.loc[idx, 'MB'] = row['MB']
    df_scanner.loc[idx, 'BF'] = row['BF']
            
shape = df_scanner.shape
# Write dataframe to excel file
sheet_name='linear'
df_scanner.to_excel(filepath_scanner, sheet_name=sheet_name)
# Format excel file
writer = pd.ExcelWriter(filepath_scanner, engine="openpyxl", mode="a")
workbook  = writer.book
sheet = workbook[sheet_name]
# Add filter
sheet.auto_filter.ref  = sheet.dimensions
# Highligt missing dicom dates
red_fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000', fill_type='solid')
green_fill = PatternFill(start_color='009900',end_color='009900', fill_type='solid')
for r in range(0,shape[0]):
    for c in range(0,shape[1]):
        if df_scanner_color.iloc[r,c] == 'RED':
            cell = sheet.cell(r+2, c+2)
            cell.fill = red_fill
        if df_scanner_color.iloc[r,c] == 'GREEN':
            cell = sheet.cell(r+2, c+2)
            cell.fill = green_fill

# Highlight columns
df_scanner_cols = list(df_scanner.columns)
sheet = highlight_columns(sheet, columns=[df_scanner_cols.index(col) for col in df_scanner_cols] , color='A5A5A5')

# # Highlight border
NumRows = df_scanner.shape[0]
NumColumns = len(columns_study)
for c in columns_study:
    col = list(df_scanner.columns).index(c) + 2
    for r in range(1, NumRows + 1):
        cell = sheet.cell(r, col)
        cell.border = Border(left=Side(style='thin'))
        
writer.save()

# Split scanner list
splitScannerList(filepath_scanner)


# from scipy.ndimage.measurements import label    

# arr = np.zeros((100,100,100))
# arr[10:20,10:20,10:20] = np.ones((10,10,10))
# arr[50:60,50:60,50:60] = np.ones((10,10,10))
# structure = np.zeros((3,3,3))
# structure[1,1,1] = 1
# structure[2,1,1] = 1
# structure[1,2,1] = 1
# structure[1,1,2] = 1
# structure[0,1,1] = 1
# structure[1,0,1] = 1
# structure[1,1,0] = 1
# labeled, ncomponents = label(arr, structure)

# import SimpleITK as sitk
# from SimpleITK import ConnectedComponentImageFilter

# comp = ConnectedComponentImageFilter()
# comp.Execute