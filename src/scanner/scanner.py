# -*- coding: utf-8 -*-

import sys, os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
from collections import defaultdict
from scanner_map import searchKey, CertifiedManufacturerModelNameCTDict, CertifiedManufacturerCTDict, TrueManufacturerModelNameCTDict, TrueManufacturerCTDict
from scanner_map import ScannerType, CertifiedManufacturerModelNameICADict, CertifiedManufacturerICADict, TrueManufacturerModelNameICADict, TrueManufacturerICADict

def highlight_columns(sheet, columns=[], color='A5A5A5', offset=2):
    for col in columns:
        cell = sheet.cell(1, col+offset)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type = 'solid')
    return sheet  

# Read discharge data
filepath = 'H:/cloud/cloud_data/Projects/CACSFilter/data/scanner/discharge_all_10062020.xlsx'
filepath_scanner = 'H:/cloud/cloud_data/Projects/DISCHARGEMaster/src/scanner/data/scanner_13022021.xlsx'
df_discharge = pd.read_excel(filepath, 'linear')
df_discharge.replace(to_replace=[np.nan], value='', inplace=True)

columns = ['Site', 'PatientID', 'StudyInstanceUID', 'Modality', 'CertifiedManufacturer', 'CertifiedManufacturerModelName','CertifiedScanner',
           'TrueManufacturer', 'TrueManufacturerModelName', 'TrueScanner']
df_scanner = pd.DataFrame(columns=columns)
# Extract list of all StudyInstanceUIDs
StudyInstanceUIDList = df_discharge['StudyInstanceUID'].copy().unique()

# Extract scnner information
for index, study in enumerate(StudyInstanceUIDList):
    print('index', index)
    df_study = df_discharge[df_discharge['StudyInstanceUID']==study]
    site = df_study['Site'].iloc[0]
    if study=='1.2.840.113619.6.95.31.0.3.4.1.1018.13.11319894':
        sys.exit()
    
    modalities = list(df_study['Modality'].unique())
    modalities_str = ','.join(modalities)

    if 'CT' in modalities_str:
        CertifiedManufacturer = CertifiedManufacturerCTDict[site]
        CertifiedManufacturerModelName = CertifiedManufacturerModelNameCTDict[site]
        CertifiedScanner = CertifiedManufacturer + ' ' + CertifiedManufacturerModelName
    
        # Filter wrong  ManufacturerModelName and  Manufacturer
        TrueManufacturerList = list(df_study['Manufacturer'])
        TrueManufacturer_str = ','.join(list(set(TrueManufacturerList)))
        TrueManufacturer = searchKey(TrueManufacturerCTDict, TrueManufacturer_str)
        TrueManufacturerModelNameList = list(df_study['ManufacturerModelName'])
        TrueManufacturerModelName_str = ','.join(list(set(TrueManufacturerModelNameList)))
        TrueManufacturerModelName = searchKey(TrueManufacturerModelNameCTDict, TrueManufacturerModelName_str) 
        TrueScanner = TrueManufacturer + ' ' + TrueManufacturerModelName
    else:
        CertifiedManufacturer = CertifiedManufacturerICADict[site]
        CertifiedManufacturerModelName = CertifiedManufacturerModelNameICADict[site]
        CertifiedScanner = CertifiedManufacturer + ' ' + CertifiedManufacturerModelName
        
        # Filter wrong  ManufacturerModelName and  Manufacturer
        TrueManufacturerList = list(df_study['Manufacturer'])
        TrueManufacturer_str = ','.join(list(set(TrueManufacturerList)))
        TrueManufacturer = searchKey(TrueManufacturerICADict, TrueManufacturer_str)
        TrueManufacturerModelNameList = list(df_study['ManufacturerModelName'])
        TrueManufacturerModelName_str = ','.join(list(set(TrueManufacturerModelNameList)))
        TrueManufacturerModelName = searchKey(TrueManufacturerModelNameICADict, TrueManufacturerModelName_str) 
        TrueScanner = TrueManufacturer + ' ' + TrueManufacturerModelName
        
    TrueScannerType = ScannerType[TrueManufacturerModelName]

    row = {'Site': df_study['Site'].iloc[0], 
           'PatientID': df_study['PatientID'].iloc[0], 
           'StudyInstanceUID': study,
           'Modality': modalities_str,
           'CertifiedManufacturer': CertifiedManufacturer,
           'CertifiedManufacturerModelName': CertifiedManufacturerModelName,
           'CertifiedScanner': CertifiedScanner,
           'TrueManufacturer': TrueManufacturer,
           #'TrueManufacturer': TrueManufacturer_str,
           'TrueManufacturerModelName': TrueManufacturerModelName,
           #'TrueManufacturerModelName': TrueManufacturerModelName_str,
           'TrueScanner': TrueScanner,
           'TrueScannerType': TrueScannerType}
    
    df_scanner = df_scanner.append(row, ignore_index=True)

sheet_name='linear'
df_scanner.to_excel(filepath_scanner, sheet_name=sheet_name)

# Format excel file
writer = pd.ExcelWriter(filepath_scanner, engine="openpyxl", mode="a")
workbook  = writer.book
sheet = workbook[sheet_name]
# Add filter
sheet.auto_filter.ref  = sheet.dimensions
# Highlight columns
df_scanner_cols = list(df_scanner.columns)
sheet = highlight_columns(sheet, columns=[df_scanner_cols.index(col) for col in df_scanner_cols] , color='A5A5A5')
writer.save() 

#TrueManufacturerModelName = df_scanner['TrueManufacturerModelName'].unique()
#TrueManufacturer = df_scanner['TrueManufacturer'].unique()

